import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"i:/Mark/my-kb/wiki/金融投資/主動型ETF持股明細.xlsx"

def add_difference_to_weekly_reductions():
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if 'Weekly Reductions' not in wb.sheetnames:
        print('Weekly Reductions sheet not found.')
        return
    ws = wb['Weekly Reductions']
    # Determine column indices (1-based). Assuming columns F and G exist.
    # Find header row (first row)
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        f_idx = header.index('持有張數') + 1  # adjust if column name differs
    except ValueError:
        f_idx = 6  # fallback to column F (6th column)
    try:
        g_idx = header.index('總加碼張數') + 1  # adjust if column name differs
    except ValueError:
        g_idx = 7  # fallback to column G (7th column)
    # Insert 差異 column after G (or at the end if not enough columns)
    diff_col_idx = max(f_idx, g_idx) + 1
    ws.insert_cols(diff_col_idx)
    ws.cell(row=1, column=diff_col_idx, value='差異')
    # Compute difference for each row
    for row in ws.iter_rows(min_row=2, max_col=diff_col_idx):
        f_val = row[f_idx - 1].value
        g_val = row[g_idx - 1].value
        try:
            diff = float(g_val) - float(f_val)
        except Exception:
            diff = None
        row[diff_col_idx - 1].value = diff
    wb.save(EXCEL_PATH)
    wb.close()
    print('差異 column added successfully.')

if __name__ == '__main__':
    add_difference_to_weekly_reductions()
