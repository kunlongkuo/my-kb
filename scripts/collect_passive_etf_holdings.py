#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Collect Taiwan passive ETF (市值型, 高股息) holdings and write Excel/Markdown summaries."""

from __future__ import annotations

import argparse
import html
import re
import time
import urllib.request
import sys
import io
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ensure UTF-8 output on Windows terminal
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE_URL = "https://www.moneydj.com/ETF/X/Basic/Basic0007B.xdjhtm?etfid={}.TW"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept-Encoding": "identity"}
DETAIL_FIELDS = [
    "ETF代號",
    "ETF名稱",
    "資料日期",
    "持股代號",
    "持股名稱",
    "投資比例(%)",
    "持有股數",
    "持有張數",
    "來源",
]
LEGACY_XLSX_SHEETS = {"工作表1", "主動型ETF持股明細", "Sheet1"}


def today_sheet_name() -> str:
    return date.today().strftime("%Y%m%d")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    default_input = Path("wiki/金融投資/台灣ETF比較清單.xlsx")
    parser.add_argument(
        "--input-file",
        type=Path,
        default=default_input if default_input.exists() else None,
        help="Excel file containing ETF lists, e.g. wiki/金融投資/台灣ETF比較清單.xlsx",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("wiki/金融投資"),
        help="Directory to output the Excel and Markdown files.",
    )
    parser.add_argument(
        "--sheet-date",
        default=today_sheet_name(),
        help="Worksheet name for this run, defaulting to today's date in YYYYMMDD format.",
    )
    parser.add_argument(
        "--top",
        type=int,
        default=50,
        help="Number of aggregate holdings to include in the Markdown summary.",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.5,
        help="Seconds to pause between MoneyDJ requests.",
    )
    return parser.parse_args()


def format_ticker(ticker_val) -> str:
    if ticker_val is None:
        return ""
    s = str(ticker_val).strip()
    if not s:
        return ""
    if s.isdigit():
        if s.startswith("0"):
            return s
        else:
            val_int = int(s)
            if val_int < 100:
                return f"{val_int:04d}"
            elif val_int < 1000:
                return f"{val_int:05d}"
            else:
                return f"{val_int:06d}"
    return s


def read_tickers_from_excel(path: Path, sheet_name: str) -> list[tuple[str, str]]:
    """Reads tickers and their names from the specified sheet of the excel file.
    
    Returns:
        List of tuples (ticker, name)
    """
    if not path.exists():
        print(f"Error: Input file {path} not found.")
        return []
    
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in {path}.")
        return []
        
    sheet = wb[sheet_name]
    tickers = []
    
    # Iterate through rows, skipping header
    # Row 1 is header: ['證券代號', '簡稱', ...]
    for r in range(2, sheet.max_row + 1):
        ticker_val = sheet.cell(row=r, column=1).value
        name_val = sheet.cell(row=r, column=2).value
        
        ticker = format_ticker(ticker_val)
        name = str(name_val or "").strip()
        
        if ticker:
            tickers.append((ticker, name))
            
    return tickers


def to_float_or_none(value: str) -> float | None:
    cleaned = value.strip().replace(",", "")
    if not cleaned or cleaned.upper() in {"N/A", "-"}:
        return None
    return float(cleaned)


def clean_text(fragment: str) -> str:
    without_tags = re.sub(r"<.*?>", "", fragment)
    return re.sub(r"\s+", " ", html.unescape(without_tags)).strip()


def split_holding(holding: str) -> tuple[str, str]:
    code_match = re.search(
        r"\(((?:[0-9]{4,6}|[A-Z.]{1,12})\.(?:TW|US)|[0-9]{4,6})\)$",
        holding,
    )
    if not code_match:
        return "", holding
    code = code_match.group(1).replace(".TW", "")
    name = holding[: code_match.start()].strip()
    return code, name


def fetch_text(url: str, timeout: int = 20) -> str:
    request = urllib.request.Request(url, headers=HEADERS)
    data = urllib.request.urlopen(request, timeout=timeout).read()
    return data.decode("utf-8", "replace")


def parse_moneydj(ticker: str) -> tuple[list[dict[str, object]], tuple[str, str, int, str]]:
    url = BASE_URL.format(ticker)
    text = fetch_text(url)

    title_match = re.search(r'<div class="eTitle">\s*(.*?)\s*<div', text, re.S)
    title = clean_text(title_match.group(1)) if title_match else ticker
    etf_name = title.replace(f"({ticker}.TW)-全部持股", "").strip()

    date_match = re.search(r"資料日期：([0-9/]+)", text)
    data_date = date_match.group(1) if date_match else ""

    matches = re.findall(
        r'<td class="col05">\s*(?:<a [^>]*>)?(.*?)(?:</a>)?\s*</td>\s*'
        r'<td class="col06">\s*([^<]*)\s*</td>\s*'
        r'<td class="col07">\s*([^<]*)\s*</td>',
        text,
        re.S,
    )

    rows: list[dict[str, object]] = []
    for holding_html, ratio_text, shares_text in matches:
        holding = clean_text(holding_html)
        holding_code, holding_name = split_holding(holding)
        ratio = to_float_or_none(ratio_text) or 0.0
        shares = to_float_or_none(shares_text)
        lots = round(shares / 1000, 3) if shares is not None else ""
        rows.append(
            {
                "ETF代號": ticker,
                "ETF名稱": etf_name,
                "資料日期": data_date,
                "持股代號": holding_code,
                "持股名稱": holding_name,
                "投資比例(%)": ratio,
                "持有股數": int(shares)
                if shares is not None and shares.is_integer()
                else (shares if shares is not None else ""),
                "持有張數": lots,
                "來源": url,
            }
        )

    return rows, (ticker, data_date or "無日期", len(matches), url)


def aggregate(rows: list[dict[str, object]]) -> list[tuple[str, dict[str, object]]]:
    holdings = defaultdict(
        lambda: {
            "name": "",
            "etfs": set(),
            "lots": 0.0,
            "lots_missing": False,
            "weight_sum": 0.0,
            "max_weight": 0.0,
            "max_etf": "",
        }
    )
    for row in rows:
        key = str(row["持股代號"] or row["持股名稱"])
        item = holdings[key]
        item["name"] = row["持股名稱"]
        item["etfs"].add(row["ETF代號"])
        if row["持有張數"] == "":
            item["lots_missing"] = True
        else:
            item["lots"] += float(row["持有張數"])
        item["weight_sum"] += float(row["投資比例(%)"])
        if float(row["投資比例(%)"]) > float(item["max_weight"]):
            item["max_weight"] = row["投資比例(%)"]
            item["max_etf"] = row["ETF代號"]

    return sorted(
        holdings.items(),
        key=lambda pair: (-float(pair[1]["weight_sum"]), -float(pair[1]["lots"]), pair[0]),
    )


def normalize_sheet_name(name: str) -> str:
    name = re.sub(r"^(\d{4})-(\d{2})-(\d{2})$", r"\1\2\3", name.strip())
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", name.strip())[:31]
    return cleaned or today_sheet_name()


def update_comparison_sheets(workbook, normalized_name: str, rows: list[dict[str, object]]) -> None:
    # Find chronological list of date sheets in the workbook (excluding normalized_name)
    existing_date_sheets = sorted(
        title for title in workbook.sheetnames 
        if re.fullmatch(r"\d{8}", title) and title != normalized_name
    )
    if not existing_date_sheets:
        print("No previous date sheets found. Skipping comparison sheets update.")
        return
        
    prev_sheet = existing_date_sheets[-1]
    print(f"Comparing '{normalized_name}' with previous sheet '{prev_sheet}' for difference sheets.")
    
    # Read previous rows
    ws_prev = workbook[prev_sheet]
    prev_rows_raw = list(ws_prev.iter_rows(values_only=True))
    if not prev_rows_raw or len(prev_rows_raw) < 2:
        return
    
    headers_prev = [str(v or "") for v in prev_rows_raw[0]]
    prev_rows = []
    for r in prev_rows_raw[1:]:
        prev_rows.append({headers_prev[i]: r[i] for i in range(len(headers_prev)) if i < len(r)})
        
    # Index by etf and holding code
    def get_indexed(row_list):
        indexed = defaultdict(dict)
        for r in row_list:
            etf_code = str(r.get("ETF代號") or "").strip()
            stock_code = str(r.get("持股代號") or "").strip()
            if etf_code and stock_code:
                indexed[etf_code][stock_code] = r
        return indexed
        
    indexed_prev = get_indexed(prev_rows)
    indexed_new = get_indexed(rows)
    
    # Format date
    formatted_date = f"{normalized_name[:4]}/{normalized_name[4:6]}/{normalized_name[6:]}"
    
    # Setup data structures
    added_list = []
    removed_list = []
    increased_list = []
    decreased_list = []
    
    # All ETFs across both sheets
    all_etfs = sorted(set(indexed_prev.keys()) | set(indexed_new.keys()))
    
    for etf_code in all_etfs:
        prev_stocks = indexed_prev.get(etf_code, {})
        new_stocks = indexed_new.get(etf_code, {})
        
        # Added (present in new, not in prev)
        for s_code, row_new in new_stocks.items():
            if s_code not in prev_stocks:
                added_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_new["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_new["持股名稱"],
                    "amount": row_new["持有張數"]
                })
                
        # Removed (present in prev, not in new)
        for s_code, row_prev in prev_stocks.items():
            if s_code not in new_stocks:
                removed_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_prev["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_prev["持股名稱"],
                    "amount": row_prev["持有張數"]
                })
                
        # Both present
        for s_code in sorted(set(prev_stocks.keys()) & set(new_stocks.keys())):
            row_prev = prev_stocks[s_code]
            row_new = new_stocks[s_code]
            
            try:
                lots_prev = float(row_prev["持有張數"]) if row_prev["持有張數"] not in (None, "") else 0.0
            except (ValueError, TypeError):
                lots_prev = 0.0
            try:
                lots_new = float(row_new["持有張數"]) if row_new["持有張數"] not in (None, "") else 0.0
            except (ValueError, TypeError):
                lots_new = 0.0
                
            if lots_new > lots_prev:
                increased_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_new["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_new["持股名稱"],
                    "lots_prev": lots_prev,
                    "lots_new": lots_new,
                    "amount": lots_new - lots_prev
                })
            elif lots_prev > lots_new:
                decreased_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_new["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_new["持股名稱"],
                    "lots_prev": lots_prev,
                    "lots_new": lots_new,
                    "amount": lots_prev - lots_new
                })

    # Styling helper for custom sheets
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="微軟正黑體", size=10)
    fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    def write_diff_sheet(sheet_title, headers, data_list):
        is_add_dec_sheet = sheet_title in ("加碼張數", "減碼張數")
        
        if sheet_title in workbook.sheetnames:
            ws_diff = workbook[sheet_title]
            
            # Read existing rows (skipping header, filter current date to avoid duplicates if re-run on the same day)
            existing_rows = []
            for r in ws_diff.iter_rows(values_only=True):
                if not existing_rows:
                    existing_rows.append(headers)
                    continue
                if all(v is None for v in r):
                    continue
                if str(r[0]).strip() == formatted_date:
                    continue
                existing_rows.append(r)
                
            ws_diff.delete_rows(1, ws_diff.max_row)
        else:
            ws_diff = workbook.create_sheet(sheet_title)
            existing_rows = [headers]
            
        next_row = 1
        for r_data in existing_rows:
            for col_idx, val in enumerate(r_data, 1):
                cell = ws_diff.cell(row=next_row, column=col_idx, value=val)
                if next_row == 1:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = font_regular
                    if col_idx in (1, 2, 4):
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx in (3, 5):
                        cell.alignment = Alignment(horizontal="left")
                    elif col_idx in (6, 7, 8):
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(val, (int, float)):
                            if isinstance(val, float) and not val.is_integer():
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'
                    elif col_idx == 9 and is_add_dec_sheet:
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(val, (int, float)):
                            cell.number_format = '0.00%'
                cell.border = thin_border
            if next_row == 1:
                ws_diff.row_dimensions[next_row].height = 25
            else:
                ws_diff.row_dimensions[next_row].height = 20
            next_row += 1
            
        # Write new differences
        for item in data_list:
            ws_diff.cell(row=next_row, column=1, value=item["date"]).alignment = Alignment(horizontal="center")
            ws_diff.cell(row=next_row, column=2, value=item["etf_code"]).alignment = Alignment(horizontal="center")
            ws_diff.cell(row=next_row, column=3, value=item["etf_name"]).alignment = Alignment(horizontal="left")
            ws_diff.cell(row=next_row, column=4, value=item["stock_code"]).alignment = Alignment(horizontal="center")
            ws_diff.cell(row=next_row, column=5, value=item["stock_name"]).alignment = Alignment(horizontal="left")
            
            if is_add_dec_sheet:
                cell_prev = ws_diff.cell(row=next_row, column=6, value=item["lots_prev"])
                cell_prev.alignment = Alignment(horizontal="right")
                if isinstance(item["lots_prev"], float) and not item["lots_prev"].is_integer():
                    cell_prev.number_format = '#,##0.00'
                else:
                    cell_prev.number_format = '#,##0'
                
                cell_new = ws_diff.cell(row=next_row, column=7, value=item["lots_new"])
                cell_new.alignment = Alignment(horizontal="right")
                if isinstance(item["lots_new"], float) and not item["lots_new"].is_integer():
                    cell_new.number_format = '#,##0.00'
                else:
                    cell_new.number_format = '#,##0'
                
                if sheet_title == "加碼張數":
                    formula_diff = f"=G{next_row}-F{next_row}"
                else:
                    formula_diff = f"=F{next_row}-G{next_row}"
                cell_amount = ws_diff.cell(row=next_row, column=8, value=formula_diff)
                cell_amount.alignment = Alignment(horizontal="right")
                cell_amount.number_format = '#,##0.00'
                
                formula_pct = f"=IF(F{next_row}=0,0,H{next_row}/F{next_row})"
                cell_pct = ws_diff.cell(row=next_row, column=9, value=formula_pct)
                cell_pct.alignment = Alignment(horizontal="right")
                cell_pct.number_format = '0.00%'
                
                max_cols = 9
            else:
                cell_amount = ws_diff.cell(row=next_row, column=6, value=item["amount"])
                cell_amount.alignment = Alignment(horizontal="right")
                val = item["amount"]
                if isinstance(val, float) and not val.is_integer():
                    cell_amount.number_format = '#,##0.00'
                else:
                    cell_amount.number_format = '#,##0'
                max_cols = 6
                
            for col in range(1, max_cols + 1):
                c = ws_diff.cell(row=next_row, column=col)
                c.font = font_regular
                c.border = thin_border
            ws_diff.row_dimensions[next_row].height = 20
            next_row += 1
            
        # Adjust column widths
        if is_add_dec_sheet:
            col_widths = {"A": 15, "B": 12, "C": 25, "D": 12, "E": 20, "F": 15, "G": 15, "H": 15, "I": 15}
        else:
            col_widths = {"A": 15, "B": 12, "C": 25, "D": 12, "E": 20, "F": 15}
        for col_letter, width in col_widths.items():
            ws_diff.column_dimensions[col_letter].width = width
        ws_diff.views.sheetView[0].showGridLines = True

    write_diff_sheet("新增個股", ["新增日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "新增張數"], added_list)
    write_diff_sheet("刪除個股", ["刪除日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "刪除張數"], removed_list)
    write_diff_sheet("加碼張數", ["加碼日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "原張數", "加碼後張數", "加碼張數", "加碼比例"], increased_list)
    write_diff_sheet("減碼張數", ["減碼日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "原張數", "減碼後張數", "減碼張數", "減碼比例"], decreased_list)
    
    # Reorder sheets to put differences first
    order_sheets = ["新增個股", "刪除個股", "加碼張數", "減碼張數"]
    sheet_list = workbook._sheets
    reordered_sheets = []
    for title in order_sheets:
        for s in sheet_list:
            if s.title == title:
                reordered_sheets.append(s)
                break
    for s in sheet_list:
        if s.title not in order_sheets:
            reordered_sheets.append(s)
    workbook._sheets = reordered_sheets


def write_xlsx_sheet(path: Path, rows: list[dict[str, object]], sheet_name: str) -> str:
    normalized_name = normalize_sheet_name(sheet_name)
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
    had_date_sheets = any(re.fullmatch(r"\d{8}", title) for title in workbook.sheetnames)

    if normalized_name in workbook.sheetnames:
        del workbook[normalized_name]

    worksheet = workbook.create_sheet(normalized_name)
    worksheet.append(DETAIL_FIELDS)
    for row in rows:
        worksheet.append([row[field] for field in DETAIL_FIELDS])
        url = str(row.get("來源") or "")
        if url.startswith("http"):
            cell = worksheet.cell(row=worksheet.max_row, column=len(DETAIL_FIELDS))
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 80)

    for worksheet_to_check in list(workbook.worksheets):
        if worksheet_to_check.max_row == 1 and worksheet_to_check.max_column == 1:
            if worksheet_to_check["A1"].value is None and len(workbook.worksheets) > 1:
                del workbook[worksheet_to_check.title]
        elif worksheet_to_check.title in LEGACY_XLSX_SHEETS:
            del workbook[worksheet_to_check.title]
        elif not had_date_sheets and worksheet_to_check.title != normalized_name:
            del workbook[worksheet_to_check.title]

    # Update the four difference sheets automatically
    try:
        update_comparison_sheets(workbook, normalized_name, rows)
    except Exception as exc:
        print(f"Error updating comparison sheets: {exc}")

    workbook.save(path)
    return normalized_name


def sheet_to_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "") for value in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[index]: row[index] for index in range(len(headers))})
    return result


def holding_key(row: dict[str, object]) -> str:
    return str(row.get("持股代號") or row.get("持股名稱") or "")


def index_by_etf(rows: list[dict[str, object]]) -> dict[str, dict[str, dict[str, object]]]:
    indexed: dict[str, dict[str, dict[str, object]]] = defaultdict(dict)
    for row in rows:
        etf = str(row.get("ETF代號") or "")
        key = holding_key(row)
        if etf and key:
            indexed[etf][key] = row
    return indexed


def date_sheets(path: Path) -> list[str]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True)
    return sorted(title for title in workbook.sheetnames if re.fullmatch(r"\d{8}", title))


def to_ratio(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def format_lot_count(value: object) -> str:
    if value in (None, ""):
        return "N/A"
    lots = float(value)
    if lots.is_integer():
        return str(int(lots))
    return f"{lots:.3f}".rstrip("0").rstrip(".")


def format_holding(row: dict[str, object]) -> str:
    code = str(row.get("持股代號") or "").strip()
    name = str(row.get("持股名稱") or "").strip()
    ratio = to_ratio(row.get("投資比例(%)"))
    lot_text = format_lot_count(row.get("持有張數"))
    label = f"{code} {name}".strip()
    return f"{label}（{ratio:.2f}%，{lot_text} 張）"


def build_changes(
    old_rows: list[dict[str, object]],
    new_rows: list[dict[str, object]],
    threshold: float = 0.0,
) -> dict[str, dict[str, list[object]]]:
    old_by_etf = index_by_etf(old_rows)
    new_by_etf = index_by_etf(new_rows)
    changes: dict[str, dict[str, list[object]]] = {}
    for etf in sorted(set(old_by_etf) | set(new_by_etf)):
        old_holdings = old_by_etf.get(etf, {})
        new_holdings = new_by_etf.get(etf, {})
        added = [new_holdings[key] for key in sorted(set(new_holdings) - set(old_holdings))]
        removed = [old_holdings[key] for key in sorted(set(old_holdings) - set(new_holdings))]
        ratio_changed = []
        for key in sorted(set(old_holdings) & set(new_holdings)):
            old_ratio = to_ratio(old_holdings[key].get("投資比例(%)"))
            new_ratio = to_ratio(new_holdings[key].get("投資比例(%)"))
            delta = new_ratio - old_ratio
            if abs(delta) > threshold:
                ratio_changed.append((old_holdings[key], new_holdings[key], delta))
        changes[etf] = {
            "added": added,
            "removed": removed,
            "ratio_changed": ratio_changed,
        }
    return changes


def write_changes_markdown(path: Path, detail_path: Path, etf_type: str) -> tuple[str, str] | None:
    sheets = date_sheets(detail_path)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(f"# {etf_type} ETF 持股變動\n\n")
        handle.write(f"資料來源：`{detail_path.name}` 最近兩個日期分頁。\n\n")
        if len(sheets) < 2:
            handle.write("> [!NOTE]\n")
            handle.write("> 目前少於兩個日期分頁，尚無法比較新增、刪除與比例變動。\n")
            return None

        old_sheet, new_sheet = sheets[-2], sheets[-1]
        changes = build_changes(
            sheet_to_rows(detail_path, old_sheet),
            sheet_to_rows(detail_path, new_sheet),
        )
        handle.write(f"比較區間：`{old_sheet}` → `{new_sheet}`\n\n")
        handle.write(
            "| ETF | 新增 | 刪除 | 比例變動 |\n"
            "| :-- | --: | --: | --: |\n"
        )
        for etf, item in changes.items():
            handle.write(
                f"| {etf} | {len(item['added'])} | {len(item['removed'])} | "
                f"{len(item['ratio_changed'])} |\n"
            )

        for etf, item in changes.items():
            handle.write(f"\n## {etf}\n\n")
            handle.write("### 新增持股\n\n")
            if item["added"]:
                for row in item["added"]:
                    handle.write(f"- {format_holding(row)}\n")
            else:
                handle.write("- 無\n")

            handle.write("\n### 刪除持股\n\n")
            if item["removed"]:
                for row in item["removed"]:
                    handle.write(f"- {format_holding(row)}\n")
            else:
                handle.write("- 無\n")

            handle.write("\n### 投資比例變動\n\n")
            if item["ratio_changed"]:
                for old_row, new_row, delta in item["ratio_changed"]:
                    label = f"{new_row.get('持股代號') or ''} {new_row.get('持股名稱') or ''}".strip()
                    old_ratio = to_ratio(old_row.get("投資比例(%)"))
                    new_ratio = to_ratio(new_row.get("投資比例(%)"))
                    handle.write(f"- {label}：{old_ratio:.2f}% → {new_ratio:.2f}%（{delta:+.2f}）\n")
            else:
                handle.write("- 無\n")
        return old_sheet, new_sheet


def format_lots(item: dict[str, object]) -> str:
    lots = float(item["lots"])
    if not item["lots_missing"]:
        return f"{lots:.3f}"
    if lots == 0:
        return "N/A"
    return f"{lots:.3f} + N/A"


def write_markdown(
    path: Path,
    statuses: list[tuple[str, str, int, str]],
    aggregate_rows: list[tuple[str, dict[str, object]]],
    detail_filename: str,
    top: int,
    etf_type: str,
) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(f"# {etf_type} ETF 持股彙總\n\n")
        handle.write(
            f"資料來源：MoneyDJ ETF 持股狀況「全部持股」頁；{etf_type} ETF 依本地清單整理。\n\n"
        )
        handle.write("> [!NOTE]\n")
        handle.write(
            "> 「持有張數」以 MoneyDJ 的持有股數除以 1,000 換算；"
            "債券或期貨等標的若來源顯示 N/A，張數留空。"
            "「加總比例」是同一持股在不同 ETF 投資比例的直接加總，"
            "適合看重疊熱度，不代表全市場加權曝險。\n\n"
        )
        handle.write("## 抓取狀態\n\n")
        handle.write("| ETF | 資料日期 | 筆數 | 來源 |\n| :-- | :-- | --: | :-- |\n")
        for ticker, data_date, count, url in statuses:
            handle.write(f"| {ticker} | {data_date} | {count} | [MoneyDJ]({url}) |\n")

        handle.write(f"\n## 依持股彙總（加總比例前 {top}）\n\n")
        handle.write(
            "| 排名 | 代號 | 名稱 | 涵蓋 ETF 數 | 總持有張數 | 加總比例(%) | "
            "最高單檔比例 | 最高比例 ETF | ETF 清單 |\n"
        )
        handle.write("| --: | :-- | :-- | --: | --: | --: | --: | :-- | :-- |\n")
        for index, (code, item) in enumerate(aggregate_rows[:top], 1):
            etfs = ", ".join(sorted(item["etfs"]))
            display_code = code if code != item["name"] else ""
            handle.write(
                f"| {index} | {display_code} | {item['name']} | {len(item['etfs'])} | "
                f"{format_lots(item)} | {float(item['weight_sum']):.2f} | "
                f"{float(item['max_weight']):.2f} | {item['max_etf']} | {etfs} |\n"
            )

        handle.write("\n## 明細檔\n\n")
        handle.write(f"- [{detail_filename}]({detail_filename})\n")
        handle.write(f"- [{etf_type}ETF持股變動.md]({etf_type}ETF持股變動.md)\n")


def process_etf_type(etf_type: str, input_file: Path, output_dir: Path, sheet_date: str, top: int, sleep_time: float) -> None:
    print(f"\n========================================")
    print(f"Processing {etf_type} ETFs...")
    print(f"========================================")
    
    tickers_info = read_tickers_from_excel(input_file, etf_type)
    if not tickers_info:
        print(f"No tickers found for type {etf_type}.")
        return

    print(f"Found {len(tickers_info)} ETFs in '{etf_type}' sheet:")
    for t, n in tickers_info:
        print(f" - {t} ({n})")
        
    all_rows: list[dict[str, object]] = []
    statuses: list[tuple[str, str, int, str]] = []
    
    for idx, (ticker, name) in enumerate(tickers_info, 1):
        print(f"[{idx}/{len(tickers_info)}] Fetching {ticker} ({name})...")
        try:
            rows, status = parse_moneydj(ticker)
            all_rows.extend(rows)
            statuses.append(status)
            print(f"   Success: {status[1]}, {status[2]} holdings parsed.")
        except Exception as exc:
            print(f"   Error fetching {ticker}: {exc}")
            # Try to add a stub status so we know it failed
            statuses.append((ticker, "抓取失敗", 0, BASE_URL.format(ticker)))
            
        time.sleep(sleep_time)

    if not all_rows:
        print(f"No data fetched for {etf_type} ETFs. Skipping file write.")
        return

    # Write detailed excel
    excel_filename = f"{etf_type}ETF持股明細.xlsx"
    excel_path = output_dir / excel_filename
    print(f"Writing detailed Excel to {excel_path}...")
    normalized_sheet = write_xlsx_sheet(excel_path, all_rows, sheet_date)
    print(f"Excel sheet '{normalized_sheet}' written successfully.")

    # Write summaries
    aggregate_rows = aggregate(all_rows)
    
    summary_filename = f"{etf_type}ETF持股彙總.md"
    summary_path = output_dir / summary_filename
    print(f"Writing summary Markdown to {summary_path}...")
    write_markdown(summary_path, statuses, aggregate_rows, excel_filename, top, etf_type)
    
    changes_filename = f"{etf_type}ETF持股變動.md"
    changes_path = output_dir / changes_filename
    print(f"Writing changes Markdown to {changes_path}...")
    write_changes_markdown(changes_path, excel_path, etf_type)
    
    print(f"Finished processing {etf_type} ETFs.")


def main() -> None:
    args = parse_args()
    if not args.input_file or not args.input_file.exists():
        print(f"Error: Input comparison list not found. Checked default path: wiki/金融投資/台灣ETF比較清單.xlsx")
        sys.exit(1)
        
    args.output_dir.mkdir(parents=True, exist_ok=True)
    
    # Process both types
    process_etf_type("市值型", args.input_file, args.output_dir, args.sheet_date, args.top, args.sleep)
    process_etf_type("高股息", args.input_file, args.output_dir, args.sheet_date, args.top, args.sleep)
    
    print("\nAll tasks completed successfully!")


if __name__ == "__main__":
    main()
