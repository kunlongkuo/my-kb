#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Sync Excel sheets from 台灣ETF比較清單.xlsx to their corresponding markdown files."""

from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook

# Sheet name to Markdown file path mapping
MAPPING = {
    "主動型": "wiki/金融投資/主動型ETF清單.md",
    "市值型": "wiki/金融投資/市值型ETF清單.md",
    "高股息": "wiki/金融投資/高股息ETF清單.md",
    "海外型": "wiki/金融投資/海外型ETF清單.md",
    "槓桿反向": "wiki/金融投資/槓桿反向型ETF清單.md",
}

def format_ticker_for_md(ticker: str, sheet_name: str) -> str:
    """Format ticker code for markdown table column 1."""
    ticker_str = str(ticker).strip()
    if sheet_name in ("市值型", "高股息"):
        return f"**'{ticker_str}**"
    return f"**{ticker_str}**"

def excel_row_to_md(row_cells, sheet_name: str) -> str:
    """Convert a row of cells to a markdown table row."""
    vals = []
    for idx, cell in enumerate(row_cells):
        val = cell.value
        if val is None:
            val_str = ""
        else:
            val_str = str(val).strip()
            
        if idx == 0:
            val_str = format_ticker_for_md(val_str, sheet_name)
            
        vals.append(val_str)
        
    return "| " + " | ".join(vals) + " |"

def sync_sheet_to_md(sheet_name: str, md_path: Path, wb) -> None:
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found in Excel.")
        return
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows:
        print(f"Warning: Sheet '{sheet_name}' is empty.")
        return
        
    # Generate the new markdown table
    headers = [str(cell.value or "").strip() for cell in rows[0]]
    header_line = "| " + " | ".join(headers) + " |"
    
    # Alignment: Column 1-5 left/center, numbers right? Let's just follow the original styles:
    # 市值型 original separator: | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |
    # 主動型 original separator: | :--------- | :-------- | :-- | :----- | :------------ | :---------- | :----------- | :----------- |
    # We can just use standard left-alignment :---
    separator_line = "| " + " | ".join([":---"] * len(headers)) + " |"
    
    table_lines = [header_line, separator_line]
    for row in rows[1:]:
        if all(cell.value is None for cell in row):
            continue
        table_lines.append(excel_row_to_md(row, sheet_name))
        
    new_table_str = "\n".join(table_lines)
    
    # Read the markdown file
    if not md_path.exists():
        print(f"Warning: Markdown file '{md_path}' does not exist. Skipping.")
        return
        
    content = md_path.read_text(encoding="utf-8")
    
    # Locate the table using regex: starts with a line containing `| 證券代號 |` or similar, 
    # and includes all lines starting with `|`.
    # Let's search for the first line starting with `| 證券代號` or `|證券代號`
    pattern = r"(\|\s*證券代號\s*\|.*?\n(?:\|.*?\n)*)"
    
    match = re.search(pattern, content, re.S)
    if not match:
        # Try a more relaxed check (just the first line starting with | 證券代號)
        pattern_relaxed = r"(\|.*?證券代號.*?\n(?:\|.*?\n?)*)"
        match = re.search(pattern_relaxed, content, re.S)
        
    if match:
        old_table = match.group(1).rstrip()
        # Keep trailing newlines
        content = content.replace(old_table, new_table_str)
        md_path.write_text(content, encoding="utf-8", newline="\n")
        print(f"Successfully updated '{md_path.name}' table.")
    else:
        print(f"Error: Could not find table starting with '| 證券代號' in '{md_path.name}'.")

def main() -> None:
    xlsx_path = Path("wiki/金融投資/台灣ETF比較清單.xlsx")
    if not xlsx_path.exists():
        print(f"Error: Excel file '{xlsx_path}' does not exist.")
        return
        
    wb = load_workbook(xlsx_path, data_only=True)
    for sheet_name, md_file in MAPPING.items():
        md_path = Path(md_file)
        sync_sheet_to_md(sheet_name, md_path, wb)

if __name__ == "__main__":
    main()
