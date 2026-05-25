import openpyxl, sys

EXCEL_PATH = r"i:/Mark/my-kb/wiki/金融投資/主動型ETF持股明細.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Weekly Reductions']

# Header
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print('Header:', header)

zero_rows = []
red_rows = []
for row in ws.iter_rows(min_row=2, values_only=False):
    # Assuming column names include '持有張數' and '持股名稱'
    # Map header to cell values
    row_dict = {header[i]: cell.value for i, cell in enumerate(row)}
    if row_dict.get('持有張數') == 0:
        zero_rows.append(row_dict)
    # Check any cell red font
    if any(cell.font and cell.font.color and cell.font.color.rgb == 'FF0000' for cell in row):
        red_rows.append(row_dict)

print('Rows with zero holding count:', len(zero_rows))
if zero_rows:
    print('Sample zero rows (first 5):')
    for r in zero_rows[:5]:
        print(r)

print('Rows with red font count:', len(red_rows))
if red_rows:
    print('Sample red rows (first 5):')
    for r in red_rows[:5]:
        print(r)
