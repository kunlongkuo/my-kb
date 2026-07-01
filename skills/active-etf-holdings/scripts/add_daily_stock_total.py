#!/usr/bin/env python
"""
add_daily_stock_total.py

讀取主動型ETF持股明細.xlsx 中所有 YYYYMMDD 日期分頁，
彙總每日所有 ETF 對每支個股的「總持有張數」，
並與前一日相比計算增減張數及增減比例，
將結果寫入同一個 XLSX 檔的「每日個股合計」分頁。

欄位說明：
  日期          - 資料日期（YYYY/MM/DD）
  持股代號       - 個股代號（若無代號則留空）
  持股名稱       - 個股名稱
  涵蓋ETF數      - 當日持有該個股的 ETF 檔數
  ETF清單        - 當日持有該個股的 ETF 代號（逗號分隔）
  當日總張數      - 當日所有 ETF 合計持有該個股的張數
  前日總張數      - 前一個日期分頁的合計張數（首日留空）
  張數增減        - 當日總張數 - 前日總張數（首日留空）
  增減比例(%)     - 張數增減 ÷ 前日總張數（存為小數，如 0.05 = 5%；前日為 0 或首日留空）

執行方式（在知識庫根目錄下）：
  python skills/active-etf-holdings/scripts/add_daily_stock_total.py

可透過 --xlsx 參數指定不同路徑的 XLSX 檔案：
  python skills/active-etf-holdings/scripts/add_daily_stock_total.py \\
      --xlsx "wiki/金融投資/主動型ETF持股明細.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

# ─────────────────────────────────────────────────────────────────────────────
# 預設路徑（在知識庫根目錄執行時有效）
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_XLSX = Path("wiki/金融投資/主動型ETF持股明細.xlsx")
OUTPUT_SHEET = "每日個股合計"

OUTPUT_HEADERS = [
    "日期",
    "持股代號",
    "持股名稱",
    "涵蓋ETF數",
    "ETF清單",
    "當日總張數",
    "前日總張數",
    "張數增減",
    "增減比例(%)",
]


# ─────────────────────────────────────────────────────────────────────────────
# 樣式
# ─────────────────────────────────────────────────────────────────────────────
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
FONT_HEADER = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="微軟正黑體", size=10)
FILL_HEADER = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
FILL_INC = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 淺綠（增加）
FILL_DEC = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 淺橘（減少）
FILL_NEW = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # 淺藍（首日或新增）


# ─────────────────────────────────────────────────────────────────────────────
# 輔助函式
# ─────────────────────────────────────────────────────────────────────────────

def is_date_sheet(name: str) -> bool:
    """判斷分頁名稱是否為 YYYYMMDD 格式。"""
    return bool(re.fullmatch(r"\d{8}", name))


def fmt_date(sheet_name: str) -> str:
    """將 20260527 轉換為 2026/05/27。"""
    return f"{sheet_name[:4]}/{sheet_name[4:6]}/{sheet_name[6:]}"


def to_float(value: object) -> float | None:
    """將欄位值轉為 float，無法轉換時回傳 None。"""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_date_sheet(ws) -> dict[str, dict]:
    """
    讀取一個 YYYYMMDD 分頁，回傳以「持股鍵」為 key 的字典。
    持股鍵 = 持股代號（若無則用持股名稱）。
    值為 {name, lots, etfs}。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(v or "").strip() for v in rows[0]]

    # 欄位索引
    try:
        idx_etf_code = headers.index("ETF代號")
        idx_stock_code = headers.index("持股代號")
        idx_stock_name = headers.index("持股名稱")
        idx_lots = headers.index("持有張數")
    except ValueError as exc:
        raise RuntimeError(f"找不到必要欄位：{exc}") from exc

    # 按持股彙總（同一分頁可能有多個 ETF 持有同一個股）
    aggregated: dict[str, dict] = {}

    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # 跳過全空列

        etf_code = str(row[idx_etf_code] or "").strip()
        stock_code = str(row[idx_stock_code] or "").strip()
        stock_name = str(row[idx_stock_name] or "").strip()
        lots_raw = row[idx_lots]

        # 持股鍵：優先用代號，無代號則用名稱
        key = stock_code if stock_code else stock_name
        if not key:
            continue

        lots = to_float(lots_raw)

        if key not in aggregated:
            aggregated[key] = {
                "code": stock_code,
                "name": stock_name,
                "lots": 0.0,
                "lots_missing": False,
                "etfs": set(),
            }

        entry = aggregated[key]
        if lots is None:
            entry["lots_missing"] = True
        else:
            entry["lots"] += lots

        if etf_code:
            entry["etfs"].add(etf_code)

    return aggregated


def build_all_days(workbook) -> list[tuple[str, dict[str, dict]]]:
    """
    讀取所有 YYYYMMDD 分頁並回傳按日期排序的清單。
    每個元素為 (sheet_name, {stock_key: {code, name, lots, etfs}})。
    """
    date_sheets = sorted(
        name for name in workbook.sheetnames if is_date_sheet(name)
    )
    result = []
    for name in date_sheets:
        ws = workbook[name]
        data = read_date_sheet(ws)
        result.append((name, data))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 寫入分頁
# ─────────────────────────────────────────────────────────────────────────────

def write_output_sheet(workbook, all_days: list[tuple[str, dict]]) -> None:
    """將彙總結果寫入「每日個股合計」分頁。"""

    # 若分頁已存在，先刪除再重建
    if OUTPUT_SHEET in workbook.sheetnames:
        del workbook[OUTPUT_SHEET]

    ws = workbook.create_sheet(OUTPUT_SHEET)

    # ── 標頭列 ──
    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[1].height = 25

    # 欄寬設定
    col_widths = {
        "A": 14,  # 日期
        "B": 12,  # 持股代號
        "C": 20,  # 持股名稱
        "D": 12,  # 涵蓋ETF數
        "E": 60,  # ETF清單
        "F": 14,  # 當日總張數
        "G": 14,  # 前日總張數
        "H": 12,  # 張數增減
        "I": 14,  # 增減比例(%)
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # ── 資料列 ──
    row_idx = 2

    for day_idx, (sheet_name, today_data) in enumerate(all_days):
        date_str = fmt_date(sheet_name)

        # 前一日資料（若有）
        prev_data: dict[str, dict] = all_days[day_idx - 1][1] if day_idx > 0 else {}

        # 取得當日所有持股鍵，依加總張數排序（大到小）
        sorted_keys = sorted(
            today_data.keys(),
            key=lambda k: -today_data[k]["lots"],
        )

        for key in sorted_keys:
            entry = today_data[key]
            today_lots = entry["lots"]
            prev_entry = prev_data.get(key)
            prev_lots = prev_entry["lots"] if prev_entry else None

            # 增減張數與比例
            if prev_lots is not None:
                delta_lots = today_lots - prev_lots
                if prev_lots != 0:
                    delta_pct = delta_lots / prev_lots  # 存小數，Excel 的 0.00% 格式自動 ×100 顯示
                else:
                    delta_pct = None  # 前日為 0 無法計算比例
            else:
                delta_lots = None
                delta_pct = None

            # 決定列底色
            if prev_lots is None:
                fill = FILL_NEW   # 首日或該個股新出現
            elif delta_lots is not None and delta_lots > 0:
                fill = FILL_INC   # 增加
            elif delta_lots is not None and delta_lots < 0:
                fill = FILL_DEC   # 減少
            else:
                fill = None       # 無變動

            etf_list = ", ".join(sorted(entry["etfs"]))

            # 寫入各欄
            def write_cell(col: int, value, h_align: str = "center", fmt: str | None = None):
                c = ws.cell(row=row_idx, column=col, value=value)
                c.font = FONT_BODY
                c.alignment = Alignment(horizontal=h_align, vertical="center")
                c.border = THIN
                if fill:
                    c.fill = fill
                if fmt:
                    c.number_format = fmt
                return c

            write_cell(1, date_str, "center")
            write_cell(2, entry["code"] or "", "center")
            write_cell(3, entry["name"], "left")
            write_cell(4, len(entry["etfs"]), "center", "#,##0")
            write_cell(5, etf_list, "left")
            write_cell(6, round(today_lots, 1), "right", "#,##0.0")
            write_cell(7, round(prev_lots, 1) if prev_lots is not None else None, "right", "#,##0.0")
            write_cell(8, round(delta_lots, 1) if delta_lots is not None else None, "right", "#,##0.0;[Red]-#,##0.0")
            write_cell(9, round(delta_pct, 4) if delta_pct is not None else None, "right", "0.00%;[Red]-0.00%")

            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

    # ── 凍結標頭 & 自動篩選 ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{row_idx - 1}"

    print(f"已寫入「{OUTPUT_SHEET}」分頁，共 {row_idx - 2} 列資料。")


def reorder_sheets(workbook) -> None:
    """
    將「每日個股合計」分頁移至所有 YYYYMMDD 分頁之前，
    但排在已有的差異分頁（新增個股、刪除個股、加碼張數、減碼張數）之後。
    """
    priority = ["新增個股", "刪除個股", "加碼張數", "減碼張數", OUTPUT_SHEET]
    sheet_list = workbook._sheets
    reordered = []

    for title in priority:
        for s in sheet_list:
            if s.title == title:
                reordered.append(s)
                break

    for s in sheet_list:
        if s.title not in priority:
            reordered.append(s)

    workbook._sheets = reordered


# ─────────────────────────────────────────────────────────────────────────────
# 主程式
# ─────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"XLSX 檔案路徑（預設：{DEFAULT_XLSX}）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path: Path = args.xlsx

    if not xlsx_path.exists():
        raise FileNotFoundError(f"找不到 XLSX 檔案：{xlsx_path}")

    print(f"讀取：{xlsx_path}")
    workbook = load_workbook(xlsx_path)

    date_sheets = [n for n in workbook.sheetnames if is_date_sheet(n)]
    if not date_sheets:
        print("未找到任何 YYYYMMDD 格式的日期分頁，程式結束。")
        return

    print(f"找到 {len(date_sheets)} 個日期分頁：{', '.join(date_sheets)}")

    all_days = build_all_days(workbook)
    write_output_sheet(workbook, all_days)
    reorder_sheets(workbook)

    workbook.save(xlsx_path)
    print(f"已儲存：{xlsx_path}")

    # ─────────────────────────────────────────────────────────────────────────────
    # 自動調用繪圖與 Markdown 嵌入腳本
    # ─────────────────────────────────────────────────────────────────────────────
    try:
        script_dir = Path(__file__).resolve().parent
        draw_script = script_dir / "draw_holdings_charts.py"
        if draw_script.exists():
            print("自動執行加減碼視覺化繪圖...")
            subprocess.run([sys.executable, str(draw_script)], check=True)
        else:
            print(f"找不到繪圖腳本: {draw_script}")
    except Exception as exc:
        print(f"執行繪圖腳本失敗: {exc}")


if __name__ == "__main__":
    main()
