#!/usr/bin/env python
"""Fill reduction values in the '減碼張數' sheet of the ETF holdings Excel.

For each row it computes:
    diff = 原張數 - 減碼後張數
    ratio = diff / 原張數   (0 if 原張數 is 0)
and writes these numbers into the corresponding columns (減碼張數 and 減碼比例).
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

def main() -> None:
    excel_path = Path(r"i:\\Mark\\my-kb\\wiki\\金融投資\\主動型ETF持股明細.xlsx")
    if not excel_path.is_file():
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    wb = load_workbook(excel_path)
    sheet_name = "減碼張數"
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook.")
        sys.exit(1)

    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    try:
        idx_original = header.index("原張數") + 1  # 1‑based column index
        idx_new = header.index("減碼後張數") + 1
    except ValueError:
        print("Required columns 原張數 or 減碼後張數 not found in header.")
        sys.exit(1)

    idx_diff = idx_new + 1   # 減碼張數 column
    idx_pct = idx_new + 2    # 減碼比例 column

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        original = row[idx_original - 1].value
        new = row[idx_new - 1].value
        orig_val = float(original) if original not in (None, "") else 0.0
        new_val = float(new) if new not in (None, "") else 0.0
        if orig_val == 0:
            diff = 0.0
            pct = 0.0
        else:
            diff = orig_val - new_val
            pct = diff / orig_val
        ws.cell(row=row[0].row, column=idx_diff, value=diff)
        ws.cell(row=row[0].row, column=idx_pct, value=pct)
        ws.cell(row=row[0].row, column=idx_diff).number_format = '#,##0.00'
        ws.cell(row=row[0].row, column=idx_pct).number_format = '0.00%'

    wb.save(excel_path)
    print(f"Successfully updated sheet '{sheet_name}' in {excel_path}")

if __name__ == "__main__":
    main()
