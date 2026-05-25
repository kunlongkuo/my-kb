#!/usr/bin/env python
"""Collect Taiwan active ETF holdings and write Excel/Markdown summaries."""

from __future__ import annotations

import argparse
import html
import re
import time
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



DEFAULT_TICKERS = [
    "00400A",
    "00401A",
    "00402A",
    "00403A",
    "00404A",
    "00405A",
    "00980A",
    "00981A",
    "00982A",
    "00983A",
    "00984A",
    "00985A",
    "00986A",
    "00987A",
    "00988A",
    "00989A",
    "00990A",
    "00991A",
    "00992A",
    "00993A",
    "00994A",
    "00995A",
    "00996A",
    "00997A",
    "00998A",
    "00999A",
    "00980D",
    "00983D",
]

BASE_URL = "https://www.moneydj.com/ETF/X/Basic/Basic0007B.xdjhtm?etfid={}.TW"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept-Encoding": "identity"}
DETAIL_FIELDS = [
    "ETF代號",
    "ETF名稱",
    "資料日期",
    "持股代號",
    "持股名稱",
    "投資比例(%)",
    "持有股數",
    "持有張數",
    "來源",
]
LEGACY_XLSX_SHEETS = {"工作表1", "主動型ETF持股明細"}


def today_sheet_name() -> str:
    return date.today().strftime("%Y%m%d")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    default_input = Path("wiki/金融投資/主動型ETF清單.md")
    parser.add_argument(
        "--input-list",
        type=Path,
        default=default_input if default_input.exists() else None,
        help="Markdown file containing ETF tickers, such as wiki/金融投資/主動型ETF清單.md.",
    )
    parser.add_argument(
        "--tickers",
        help="Comma-separated ETF tickers. Overrides --input-list when provided.",
    )
    default_output = Path("wiki/金融投資")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=default_output if default_output.exists() else Path("."),
        help="Directory for 主動型ETF持股明細.xlsx and 主動型ETF持股彙總.md.",
    )
    parser.add_argument(
        "--sheet-date",
        default=today_sheet_name(),
        help="Worksheet name for this run, defaulting to today's date in YYYYMMDD format.",
    )
    parser.add_argument(
        "--top",
        type=int,
        default=50,
        help="Number of aggregate holdings to include in the Markdown summary.",
    )
    parser.add_argument(
        "--change-threshold",
        type=float,
        default=0.0,
        help="Minimum absolute investment ratio change to report between the latest two worksheets.",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.2,
        help="Seconds to pause between MoneyDJ requests.",
    )
    return parser.parse_args()


def tickers_from_markdown(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8")
    tickers = re.findall(r"\b[0-9]{5}[AD]\b", text)
    return list(dict.fromkeys(tickers))


def resolve_tickers(args: argparse.Namespace) -> list[str]:
    if args.tickers:
        return [item.strip().upper() for item in args.tickers.split(",") if item.strip()]
    if args.input_list and args.input_list.exists():
        tickers = tickers_from_markdown(args.input_list)
        if tickers:
            return tickers
    return DEFAULT_TICKERS


def to_float_or_none(value: str) -> float | None:
    cleaned = value.strip().replace(",", "")
    if not cleaned or cleaned.upper() in {"N/A", "-"}:
        return None
    return float(cleaned)


def clean_text(fragment: str) -> str:
    without_tags = re.sub(r"<.*?>", "", fragment)
    return re.sub(r"\s+", " ", html.unescape(without_tags)).strip()


def split_holding(holding: str) -> tuple[str, str]:
    code_match = re.search(
        r"\(((?:[0-9]{4,6}|[A-Z.]{1,12})\.(?:TW|US)|[0-9]{4,6})\)$",
        holding,
    )
    if not code_match:
        return "", holding
    code = code_match.group(1).replace(".TW", "")
    name = holding[: code_match.start()].strip()
    return code, name


def fetch_text(url: str, timeout: int = 20) -> str:
    request = urllib.request.Request(url, headers=HEADERS)
    data = urllib.request.urlopen(request, timeout=timeout).read()
    return data.decode("utf-8", "replace")


def parse_moneydj(ticker: str) -> tuple[list[dict[str, object]], tuple[str, str, int, str]]:
    url = BASE_URL.format(ticker)
    text = fetch_text(url)

    title_match = re.search(r'<div class="eTitle">\s*(.*?)\s*<div', text, re.S)
    title = clean_text(title_match.group(1)) if title_match else ticker
    etf_name = title.replace(f"({ticker}.TW)-全部持股", "").strip()

    date_match = re.search(r"資料日期：([0-9/]+)", text)
    data_date = date_match.group(1) if date_match else ""

    matches = re.findall(
        r'<td class="col05">\s*(?:<a [^>]*>)?(.*?)(?:</a>)?\s*</td>\s*'
        r'<td class="col06">\s*([^<]*)\s*</td>\s*'
        r'<td class="col07">\s*([^<]*)\s*</td>',
        text,
        re.S,
    )

    rows: list[dict[str, object]] = []
    for holding_html, ratio_text, shares_text in matches:
        holding = clean_text(holding_html)
        holding_code, holding_name = split_holding(holding)
        ratio = to_float_or_none(ratio_text) or 0.0
        shares = to_float_or_none(shares_text)
        lots = round(shares / 1000, 3) if shares is not None else ""
        rows.append(
            {
                "ETF代號": ticker,
                "ETF名稱": etf_name,
                "資料日期": data_date,
                "持股代號": holding_code,
                "持股名稱": holding_name,
                "投資比例(%)": ratio,
                "持有股數": int(shares)
                if shares is not None and shares.is_integer()
                else (shares if shares is not None else ""),
                "持有張數": lots,
                "來源": url,
            }
        )

    return rows, (ticker, data_date or "無日期", len(matches), url)


def aggregate(rows: list[dict[str, object]]) -> list[tuple[str, dict[str, object]]]:
    holdings = defaultdict(
        lambda: {
            "name": "",
            "etfs": set(),
            "lots": 0.0,
            "lots_missing": False,
            "weight_sum": 0.0,
            "max_weight": 0.0,
            "max_etf": "",
        }
    )
    for row in rows:
        key = str(row["持股代號"] or row["持股名稱"])
        item = holdings[key]
        item["name"] = row["持股名稱"]
        item["etfs"].add(row["ETF代號"])
        if row["持有張數"] == "":
            item["lots_missing"] = True
        else:
            item["lots"] += float(row["持有張數"])
        item["weight_sum"] += float(row["投資比例(%)"])
        if float(row["投資比例(%)"]) > float(item["max_weight"]):
            item["max_weight"] = row["投資比例(%)"]
            item["max_etf"] = row["ETF代號"]

    return sorted(
        holdings.items(),
        key=lambda pair: (-float(pair[1]["weight_sum"]), -float(pair[1]["lots"]), pair[0]),
    )


def normalize_sheet_name(name: str) -> str:
    name = re.sub(r"^(\d{4})-(\d{2})-(\d{2})$", r"\1\2\3", name.strip())
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", name.strip())[:31]
    return cleaned or today_sheet_name()


def update_comparison_sheets(workbook, normalized_name: str, rows: list[dict[str, object]]) -> None:
    # Find chronological list of date sheets in the workbook (excluding normalized_name)
    existing_date_sheets = sorted(
        title for title in workbook.sheetnames 
        if re.fullmatch(r"\d{8}", title) and title != normalized_name
    )
    if not existing_date_sheets:
        print("No previous date sheets found. Skipping comparison sheets update.")
        return
        
    prev_sheet = existing_date_sheets[-1]
    print(f"Comparing '{normalized_name}' with previous sheet '{prev_sheet}' for difference sheets.")
    
    # Read previous rows
    ws_prev = workbook[prev_sheet]
    prev_rows_raw = list(ws_prev.iter_rows(values_only=True))
    if not prev_rows_raw or len(prev_rows_raw) < 2:
        return
    
    headers_prev = [str(v or "") for v in prev_rows_raw[0]]
    prev_rows = []
    for r in prev_rows_raw[1:]:
        prev_rows.append({headers_prev[i]: r[i] for i in range(len(headers_prev)) if i < len(r)})
        
    # Index by etf and holding code
    def get_indexed(row_list):
        indexed = defaultdict(dict)
        for r in row_list:
            etf_code = str(r.get("ETF代號") or "").strip()
            stock_code = str(r.get("持股代號") or "").strip()
            if etf_code and stock_code:
                indexed[etf_code][stock_code] = r
        return indexed
        
    indexed_prev = get_indexed(prev_rows)
    indexed_new = get_indexed(rows)
    
    # Format date
    formatted_date = f"{normalized_name[:4]}/{normalized_name[4:6]}/{normalized_name[6:]}"
    
    # Setup data structures
    added_list = []
    removed_list = []
    increased_list = []
    decreased_list = []
    
    # All ETFs across both sheets
    all_etfs = sorted(set(indexed_prev.keys()) | set(indexed_new.keys()))
    
    for etf_code in all_etfs:
        prev_stocks = indexed_prev.get(etf_code, {})
        new_stocks = indexed_new.get(etf_code, {})
        
        # Added (present in new, not in prev)
        for s_code, row_new in new_stocks.items():
            if s_code not in prev_stocks:
                added_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_new["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_new["持股名稱"],
                    "amount": row_new["持有張數"]
                })
                
        # Removed (present in prev, not in new)
        for s_code, row_prev in prev_stocks.items():
            if s_code not in new_stocks:
                removed_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_prev["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_prev["持股名稱"],
                    "amount": row_prev["持有張數"]
                })
                
        # Both present
        for s_code in sorted(set(prev_stocks.keys()) & set(new_stocks.keys())):
            row_prev = prev_stocks[s_code]
            row_new = new_stocks[s_code]
            
            try:
                lots_prev = float(row_prev["持有張數"]) if row_prev["持有張數"] not in (None, "") else 0.0
            except (ValueError, TypeError):
                lots_prev = 0.0
            try:
                lots_new = float(row_new["持有張數"]) if row_new["持有張數"] not in (None, "") else 0.0
            except (ValueError, TypeError):
                lots_new = 0.0
                
            if lots_new > lots_prev:
                increased_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_new["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_new["持股名稱"],
                    "amount": lots_new - lots_prev
                })
            elif lots_prev > lots_new:
                decreased_list.append({
                    "date": formatted_date,
                    "etf_code": etf_code,
                    "etf_name": row_new["ETF名稱"],
                    "stock_code": s_code,
                    "stock_name": row_new["持股名稱"],
                    "amount": lots_prev - lots_new
                })

    # Styling helper for custom sheets
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="微軟正黑體", size=10)
    fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    def write_diff_sheet(sheet_title, headers, data_list):
        if sheet_title in workbook.sheetnames:
            ws_diff = workbook[sheet_title]
            
            # 讀取既存資料（跳過首行標頭與空白列，並過濾掉當前日期的資料以免重跑時重複）
            existing_rows = []
            for r in ws_diff.iter_rows(values_only=True):
                if not existing_rows:
                    existing_rows.append(headers)  # 確保標頭正確
                    continue
                if all(v is None for v in r):
                    continue
                # 過濾掉當天同日期的資料以免重跑重複累加
                if str(r[0]).strip() == formatted_date:
                    continue
                existing_rows.append(r)
                
            # 完全清空工作表的所有列以利重新排列
            ws_diff.delete_rows(1, ws_diff.max_row)
        else:
            ws_diff = workbook.create_sheet(sheet_title)
            existing_rows = [headers]
            
        # 寫入既存資料（含標頭）
        next_row = 1
        for r_data in existing_rows:
            for col_idx, val in enumerate(r_data, 1):
                cell = ws_diff.cell(row=next_row, column=col_idx, value=val)
                if next_row == 1:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = font_regular
                    if col_idx in (1, 2, 4):
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx in (3, 5):
                        cell.alignment = Alignment(horizontal="left")
                    elif col_idx == 6:
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(val, float) and not val.is_integer():
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                cell.border = thin_border
            if next_row == 1:
                ws_diff.row_dimensions[next_row].height = 25
            else:
                ws_diff.row_dimensions[next_row].height = 20
            next_row += 1
            
        # 寫入本次新變動的資料
        for item in data_list:
            ws_diff.cell(row=next_row, column=1, value=item["date"]).alignment = Alignment(horizontal="center")
            ws_diff.cell(row=next_row, column=2, value=item["etf_code"]).alignment = Alignment(horizontal="center")
            ws_diff.cell(row=next_row, column=3, value=item["etf_name"]).alignment = Alignment(horizontal="left")
            ws_diff.cell(row=next_row, column=4, value=item["stock_code"]).alignment = Alignment(horizontal="center")
            ws_diff.cell(row=next_row, column=5, value=item["stock_name"]).alignment = Alignment(horizontal="left")
            
            cell_amount = ws_diff.cell(row=next_row, column=6, value=item["amount"])
            cell_amount.alignment = Alignment(horizontal="right")
            
            # 格式化數值
            val = item["amount"]
            if isinstance(val, float) and not val.is_integer():
                cell_amount.number_format = '#,##0.00'
            else:
                cell_amount.number_format = '#,##0'
                
            for col in range(1, 7):
                c = ws_diff.cell(row=next_row, column=col)
                c.font = font_regular
                c.border = thin_border
            ws_diff.row_dimensions[next_row].height = 20
            next_row += 1
            
        # 欄位寬度調整
        col_widths = {"A": 15, "B": 12, "C": 25, "D": 12, "E": 20, "F": 15}
        for col_letter, width in col_widths.items():
            ws_diff.column_dimensions[col_letter].width = width
        ws_diff.views.sheetView[0].showGridLines = True

    write_diff_sheet("新增個股", ["新增日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "新增張數"], added_list)
    write_diff_sheet("刪除個股", ["刪除日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "刪除張數"], removed_list)
    write_diff_sheet("加碼張數", ["加碼日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "加碼張數"], increased_list)
    write_diff_sheet("減碼張數", ["減碼日期", "ETF代號", "ETF名稱", "個股代號", "個股名稱", "減碼張數"], decreased_list)
    
    # Reorder sheets so the comparison sheets are at the very beginning (indexes 0, 1, 2, 3)
    order_sheets = ["新增個股", "刪除個股", "加碼張數", "減碼張數"]
    sheet_list = workbook._sheets
    reordered_sheets = []
    for title in order_sheets:
        for s in sheet_list:
            if s.title == title:
                reordered_sheets.append(s)
                break
    for s in sheet_list:
        if s.title not in order_sheets:
            reordered_sheets.append(s)
    workbook._sheets = reordered_sheets


def write_xlsx_sheet(path: Path, rows: list[dict[str, object]], sheet_name: str) -> str:
    normalized_name = normalize_sheet_name(sheet_name)
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
    had_date_sheets = any(re.fullmatch(r"\d{8}", title) for title in workbook.sheetnames)

    if normalized_name in workbook.sheetnames:
        del workbook[normalized_name]

    worksheet = workbook.create_sheet(normalized_name)
    worksheet.append(DETAIL_FIELDS)
    for row in rows:
        worksheet.append([row[field] for field in DETAIL_FIELDS])
        url = str(row.get("來源") or "")
        if url.startswith("http"):
            cell = worksheet.cell(row=worksheet.max_row, column=len(DETAIL_FIELDS))
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 80)

    for worksheet_to_check in list(workbook.worksheets):
        if worksheet_to_check.max_row == 1 and worksheet_to_check.max_column == 1:
            if worksheet_to_check["A1"].value is None and len(workbook.worksheets) > 1:
                del workbook[worksheet_to_check.title]
        elif worksheet_to_check.title in LEGACY_XLSX_SHEETS:
            del workbook[worksheet_to_check.title]
        elif not had_date_sheets and worksheet_to_check.title != normalized_name:
            del workbook[worksheet_to_check.title]

    # Automatically compute and write the four difference comparison sheets
    try:
        update_comparison_sheets(workbook, normalized_name, rows)
    except Exception as exc:
        print(f"Error updating comparison sheets: {exc}")

    workbook.save(path)
    return normalized_name



def sheet_to_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "") for value in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[index]: row[index] for index in range(len(headers))})
    return result


def holding_key(row: dict[str, object]) -> str:
    return str(row.get("持股代號") or row.get("持股名稱") or "")


def index_by_etf(rows: list[dict[str, object]]) -> dict[str, dict[str, dict[str, object]]]:
    indexed: dict[str, dict[str, dict[str, object]]] = defaultdict(dict)
    for row in rows:
        etf = str(row.get("ETF代號") or "")
        key = holding_key(row)
        if etf and key:
            indexed[etf][key] = row
    return indexed


def date_sheets(path: Path) -> list[str]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True)
    return sorted(title for title in workbook.sheetnames if re.fullmatch(r"\d{8}", title))


def to_ratio(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def format_lot_count(value: object) -> str:
    if value in (None, ""):
        return "N/A"
    lots = float(value)
    if lots.is_integer():
        return str(int(lots))
    return f"{lots:.3f}".rstrip("0").rstrip(".")


def format_holding(row: dict[str, object]) -> str:
    code = str(row.get("持股代號") or "").strip()
    name = str(row.get("持股名稱") or "").strip()
    ratio = to_ratio(row.get("投資比例(%)"))
    lot_text = format_lot_count(row.get("持有張數"))
    label = f"{code} {name}".strip()
    return f"{label}（{ratio:.2f}%，{lot_text} 張）"


def build_changes(
    old_rows: list[dict[str, object]],
    new_rows: list[dict[str, object]],
    threshold: float,
) -> dict[str, dict[str, list[object]]]:
    old_by_etf = index_by_etf(old_rows)
    new_by_etf = index_by_etf(new_rows)
    changes: dict[str, dict[str, list[object]]] = {}
    for etf in sorted(set(old_by_etf) | set(new_by_etf)):
        old_holdings = old_by_etf.get(etf, {})
        new_holdings = new_by_etf.get(etf, {})
        added = [new_holdings[key] for key in sorted(set(new_holdings) - set(old_holdings))]
        removed = [old_holdings[key] for key in sorted(set(old_holdings) - set(new_holdings))]
        ratio_changed = []
        for key in sorted(set(old_holdings) & set(new_holdings)):
            old_ratio = to_ratio(old_holdings[key].get("投資比例(%)"))
            new_ratio = to_ratio(new_holdings[key].get("投資比例(%)"))
            delta = new_ratio - old_ratio
            if abs(delta) > threshold:
                ratio_changed.append((old_holdings[key], new_holdings[key], delta))
        changes[etf] = {
            "added": added,
            "removed": removed,
            "ratio_changed": ratio_changed,
        }
    return changes


def write_changes_markdown(path: Path, detail_path: Path, threshold: float) -> tuple[str, str] | None:
    sheets = date_sheets(detail_path)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write("# 主動型 ETF 持股變動\n\n")
        handle.write("資料來源：`主動型ETF持股明細.xlsx` 最近兩個日期分頁。\n\n")
        if len(sheets) < 2:
            handle.write("> [!NOTE]\n")
            handle.write("> 目前少於兩個日期分頁，尚無法比較新增、刪除與比例變動。\n")
            return None

        old_sheet, new_sheet = sheets[-2], sheets[-1]
        changes = build_changes(
            sheet_to_rows(detail_path, old_sheet),
            sheet_to_rows(detail_path, new_sheet),
            threshold,
        )
        handle.write(f"比較區間：`{old_sheet}` → `{new_sheet}`\n\n")
        handle.write(
            "| ETF | 新增 | 刪除 | 比例變動 |\n"
            "| :-- | --: | --: | --: |\n"
        )
        for etf, item in changes.items():
            handle.write(
                f"| {etf} | {len(item['added'])} | {len(item['removed'])} | "
                f"{len(item['ratio_changed'])} |\n"
            )

        for etf, item in changes.items():
            handle.write(f"\n## {etf}\n\n")
            handle.write("### 新增持股\n\n")
            if item["added"]:
                for row in item["added"]:
                    handle.write(f"- {format_holding(row)}\n")
            else:
                handle.write("- 無\n")

            handle.write("\n### 刪除持股\n\n")
            if item["removed"]:
                for row in item["removed"]:
                    handle.write(f"- {format_holding(row)}\n")
            else:
                handle.write("- 無\n")

            handle.write("\n### 投資比例變動\n\n")
            if item["ratio_changed"]:
                for old_row, new_row, delta in item["ratio_changed"]:
                    label = f"{new_row.get('持股代號') or ''} {new_row.get('持股名稱') or ''}".strip()
                    old_ratio = to_ratio(old_row.get("投資比例(%)"))
                    new_ratio = to_ratio(new_row.get("投資比例(%)"))
                    handle.write(f"- {label}：{old_ratio:.2f}% → {new_ratio:.2f}%（{delta:+.2f}）\n")
            else:
                handle.write("- 無\n")
        return old_sheet, new_sheet


def format_lots(item: dict[str, object]) -> str:
    lots = float(item["lots"])
    if not item["lots_missing"]:
        return f"{lots:.3f}"
    if lots == 0:
        return "N/A"
    return f"{lots:.3f} + N/A"


def write_markdown(
    path: Path,
    statuses: list[tuple[str, str, int, str]],
    aggregate_rows: list[tuple[str, dict[str, object]]],
    detail_filename: str,
    top: int,
) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write("# 主動型 ETF 持股彙總\n\n")
        handle.write(
            "資料來源：MoneyDJ ETF 持股狀況「全部持股」頁；主動式 ETF 依本地清單整理。\n\n"
        )
        handle.write("> [!NOTE]\n")
        handle.write(
            "> 「持有張數」以 MoneyDJ 的持有股數除以 1,000 換算；"
            "債券或期貨等標的若來源顯示 N/A，張數留空。"
            "「加總比例」是同一持股在不同 ETF 投資比例的直接加總，"
            "適合看重疊熱度，不代表全市場加權曝險。\n\n"
        )
        handle.write("## 抓取狀態\n\n")
        handle.write("| ETF | 資料日期 | 筆數 | 來源 |\n| :-- | :-- | --: | :-- |\n")
        for ticker, data_date, count, url in statuses:
            handle.write(f"| {ticker} | {data_date} | {count} | [MoneyDJ]({url}) |\n")

        handle.write(f"\n## 依持股彙總（加總比例前 {top}）\n\n")
        handle.write(
            "| 排名 | 代號 | 名稱 | 涵蓋 ETF 數 | 總持有張數 | 加總比例(%) | "
            "最高單檔比例 | 最高比例 ETF | ETF 清單 |\n"
        )
        handle.write("| --: | :-- | :-- | --: | --: | --: | --: | :-- | :-- |\n")
        for index, (code, item) in enumerate(aggregate_rows[:top], 1):
            etfs = ", ".join(sorted(item["etfs"]))
            display_code = code if code != item["name"] else ""
            handle.write(
                f"| {index} | {display_code} | {item['name']} | {len(item['etfs'])} | "
                f"{format_lots(item)} | {float(item['weight_sum']):.2f} | "
                f"{float(item['max_weight']):.2f} | {item['max_etf']} | {etfs} |\n"
            )

        handle.write("\n## 明細檔\n\n")
        handle.write(f"- [{detail_filename}]({detail_filename})\n")
        handle.write("- [主動型ETF持股變動.md](主動型ETF持股變動.md)\n")


def update_comparison_xlsx(input_list_path: Path, xlsx_path: Path) -> None:
    if not input_list_path.exists() or not xlsx_path.exists():
        return
    text = input_list_path.read_text(encoding="utf-8")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    table_rows = []
    for line in lines:
        if not line.startswith("|"):
            continue
        parts = [p.strip() for p in line.split("|")[1:-1]]
        if not parts or all(p == "" for p in parts):
            continue
        if any(p.startswith(":") or p.startswith("-") for p in parts if p):
            continue
        if parts[0] == "證券代號":
            continue
        parts[0] = parts[0].replace("**", "")
        table_rows.append(parts)
    if not table_rows:
        return
    wb = load_workbook(xlsx_path)
    sheet_name = "主動型"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(["證券代號", "基金簡稱", "經理人", "保管銀行", "配息頻率 (月份)", "經理費 (年率)", "保管費 (年率)", "最近除息日期"])
    for r in table_rows:
        ws.append(r)
    wb.save(xlsx_path)
    print(f"Updated '{sheet_name}' sheet in: {xlsx_path} with {len(table_rows)} ETFs.")


def main() -> None:
    args = parse_args()
    tickers = resolve_tickers(args)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, object]] = []
    statuses: list[tuple[str, str, int, str]] = []
    for ticker in tickers:
        try:
            ticker_rows, status = parse_moneydj(ticker)
            rows.extend(ticker_rows)
            statuses.append(status)
        except Exception as exc:  # Keep the batch useful even if one ETF fails.
            statuses.append((ticker, f"抓取失敗: {type(exc).__name__}", 0, BASE_URL.format(ticker)))
        time.sleep(args.sleep)

    detail_path = args.output_dir / "主動型ETF持股明細.xlsx"
    summary_path = args.output_dir / "主動型ETF持股彙總.md"
    changes_path = args.output_dir / "主動型ETF持股變動.md"
    sheet_name = write_xlsx_sheet(detail_path, rows, args.sheet_date)
    write_markdown(summary_path, statuses, aggregate(rows), detail_path.name, args.top)
    compared_sheets = write_changes_markdown(changes_path, detail_path, args.change_threshold)

    comparison_xlsx_path = args.output_dir / "台灣ETF比較清單.xlsx"
    if args.input_list and args.input_list.exists() and comparison_xlsx_path.exists():
        try:
            update_comparison_xlsx(args.input_list, comparison_xlsx_path)
        except Exception as exc:
            print(f"Error updating comparison Excel: {exc}")

    print(f"ETF count: {len(tickers)}")
    print(f"Holding rows: {len(rows)}")
    print(f"Detail XLSX: {detail_path}")
    print(f"Worksheet: {sheet_name}")
    print(f"Summary MD: {summary_path}")
    print(f"Changes MD: {changes_path}")
    if compared_sheets:
        print(f"Compared worksheets: {compared_sheets[0]} -> {compared_sheets[1]}")


if __name__ == "__main__":
    main()
