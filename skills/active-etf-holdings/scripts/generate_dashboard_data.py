#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Generate stock-centric active ETF holdings data for the interactive dashboard."""

import os
import re
import json
import datetime
from pathlib import Path
import openpyxl
import yfinance as yf
import pandas as pd

# Define file paths
EXCEL_PATH = Path("wiki/金融投資/主動型ETF持股明細.xlsx")
OUTPUT_JS_PATH = Path("wiki/金融投資/dashboard_data.js")

def get_close_price(df, ticker, date_str):
    """Safely extract close price for a ticker on or near a target date (handling weekends/holidays)."""
    # date_str is YYYYMMDD -> convert to YYYY-MM-DD
    target_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    if df.empty:
        return None
    try:
        timestamp = pd.Timestamp(target_date)
        # Find the latest available date that is <= target_date
        available_dates = df.index[df.index <= timestamp]
        if available_dates.empty:
            available_dates = df.index
            if available_dates.empty:
                return None
            best_date = available_dates[0]
        else:
            best_date = available_dates[-1]
            
        # Handle MultiIndex vs SingleIndex columns in yfinance Output
        if isinstance(df.columns, pd.MultiIndex):
            if ticker in df.columns.levels[0]:
                val = df.loc[best_date, (ticker, "Close")]
                return float(val) if not pd.isna(val) else None
        else:
            val = df.loc[best_date, "Close"]
            return float(val) if not pd.isna(val) else None
    except Exception:
        pass
    return None

def main():
    if not EXCEL_PATH.exists():
        print(f"Error: Excel file '{EXCEL_PATH}' does not exist.")
        return

    print("Loading active ETF workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Filter YYYYMMDD sheets
    dates = []
    for sheet in wb.sheetnames:
        if re.match(r'^\d{8}$', sheet):
            dates.append(sheet)
    dates.sort()
    
    if not dates:
        print("Error: No date sheets found in Excel.")
        return
        
    print(f"Found {len(dates)} date sheets: {dates}")
    
    stock_data = {}
    etf_info = {}
    
    # 1. Parse Excel data
    print("Parsing Excel sheets...")
    for date in dates:
        ws = wb[date]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        # Skip header
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
                
            # Columns: ETF代號, ETF名稱, 資料日期, 持股代號, 持股名稱, 投資比例(%), 持有股數, 持有張數, 來源
            etf_id = str(row[0]).strip()
            etf_name = str(row[1]).strip()
            stock_id = str(row[3]).strip() if row[3] is not None else None
            stock_name = str(row[4]).strip() if row[4] is not None else ""
            ratio = row[5]
            shares = row[6]
            
            if not stock_id:
                continue
                
            # Process ETF info
            if etf_id not in etf_info:
                issuer = "未知投信"
                for kw in ["富邦", "元大", "復華", "摩根", "台新", "國泰", "群益", "統一", "兆豐", "第一金", "野村", "凱基", "中信", "保德信", "日盛"]:
                    if kw in etf_name:
                        issuer = kw + "投信"
                        break
                etf_info[etf_id] = {"name": etf_name, "issuer": issuer}
                
            # Process stock data
            if stock_id not in stock_data:
                stock_data[stock_id] = {
                    "name": stock_name,
                    "history": {}
                }
                
            if date not in stock_data[stock_id]["history"]:
                stock_data[stock_id]["history"][date] = {
                    "close_price": None,
                    "etfs": {}
                }
                
            # Safely parse numeric values
            try:
                ratio_val = float(ratio) if ratio is not None else 0.0
            except ValueError:
                ratio_val = 0.0
                
            try:
                shares_val = int(float(shares)) if shares is not None and str(shares) != "N/A" else 0
            except ValueError:
                shares_val = 0
                
            stock_data[stock_id]["history"][date]["etfs"][etf_id] = {
                "ratio": ratio_val,
                "shares": shares_val
            }

    # 2. Calculate holdings changes, additions, reductions, and exit signals
    print("Calculating holdings signals and consecutive days...")
    for stock_id, sinfo in stock_data.items():
        history = sinfo["history"]
        stock_dates = sorted(list(history.keys()))
        etf_prev_state = {} # etf_id -> {ratio, shares, signal, consecutive_days}
        
        for i, date in enumerate(stock_dates):
            current_etfs = history[date]["etfs"]
            
            # Check for ETFs that existed in the previous date but not in the current one (Exit/出清)
            if i > 0:
                for etf_id, prev_state in list(etf_prev_state.items()):
                    if etf_id not in current_etfs and prev_state["shares"] > 0:
                        # Hand-insert an Exit record at current date
                        current_etfs[etf_id] = {
                            "ratio": 0.0,
                            "shares": 0,
                            "ratio_change": -prev_state["ratio"],
                            "shares_change": -prev_state["shares"],
                            "signal": "完全出清",
                            "consecutive_days": 1
                        }
                        etf_prev_state[etf_id] = {
                            "ratio": 0.0,
                            "shares": 0,
                            "signal": "完全出清",
                            "consecutive_days": 1
                        }
            
            # Calculate metrics for current ETFs
            for etf_id, curr_info in list(current_etfs.items()):
                if curr_info.get("signal") == "完全出清":
                    continue
                    
                curr_shares = curr_info["shares"]
                curr_ratio = curr_info["ratio"]
                
                if etf_id not in etf_prev_state:
                    # First time this ETF holds this stock in this timeframe
                    if i > 0:
                        curr_info["ratio_change"] = curr_ratio
                        curr_info["shares_change"] = curr_shares
                        curr_info["signal"] = "首次買入"
                        curr_info["consecutive_days"] = 1
                    else:
                        # First date in historical record -> baseline
                        curr_info["ratio_change"] = 0.0
                        curr_info["shares_change"] = 0
                        curr_info["signal"] = "持平"
                        curr_info["consecutive_days"] = 0
                else:
                    prev_info = etf_prev_state[etf_id]
                    prev_shares = prev_info["shares"]
                    prev_ratio = prev_info["ratio"]
                    
                    shares_diff = curr_shares - prev_shares
                    ratio_diff = curr_ratio - prev_ratio
                    
                    curr_info["ratio_change"] = round(ratio_diff, 4)
                    curr_info["shares_change"] = shares_diff
                    
                    # Signal determination
                    if prev_shares == 0 and curr_shares > 0:
                        signal = "首次買入"
                    elif shares_diff > 0:
                        signal = "加碼"
                    elif shares_diff < 0:
                        signal = "減碼"
                    else:
                        signal = "持平"
                    curr_info["signal"] = signal
                    
                    # Consecutive days tracking
                    if prev_info["signal"] == signal and signal in ["加碼", "減碼"]:
                        consecutive_days = prev_info["consecutive_days"] + 1
                    else:
                        consecutive_days = 1 if signal in ["加碼", "減碼"] else 0
                    curr_info["consecutive_days"] = consecutive_days
                    
                etf_prev_state[etf_id] = {
                    "ratio": curr_ratio,
                    "shares": curr_shares,
                    "signal": curr_info["signal"],
                    "consecutive_days": curr_info["consecutive_days"]
                }

    # 3. Retrieve stock historical close prices via yfinance
    print("Preparing tickers for yfinance download...")
    all_dates = sorted(dates)
    if all_dates:
        start_date = datetime.datetime.strptime(all_dates[0], "%Y%m%d") - datetime.timedelta(days=7)
        end_date = datetime.datetime.strptime(all_dates[-1], "%Y%m%d") + datetime.timedelta(days=3)
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")
    else:
        start_str = "2026-05-01"
        end_str = "2026-07-01"

    ticker_to_stock_id = {}
    stock_id_to_ticker = {}
    tickers_to_download = []
    
    for stock_id in stock_data.keys():
        if stock_id.isdigit():
            # Taiwan stocks default to Listed (.TW)
            ticker = f"{stock_id}.TW"
            tickers_to_download.append(ticker)
            ticker_to_stock_id[ticker] = stock_id
            stock_id_to_ticker[stock_id] = ticker
        elif stock_id.endswith(".US"):
            ticker = stock_id[:-3]
            tickers_to_download.append(ticker)
            ticker_to_stock_id[ticker] = stock_id
            stock_id_to_ticker[stock_id] = ticker
        else:
            # Skip indices / futures / other complex derivatives
            pass
            
    print(f"Downloading historical close prices for {len(tickers_to_download)} tickers from {start_str} to {end_str}...")
    try:
        df = yf.download(tickers_to_download, start=start_str, end=end_str, group_by="ticker")
    except Exception as e:
        print(f"yfinance download failed: {e}")
        df = pd.DataFrame()

    # Retry missing Taiwan tickers with .TWO (Overs-The-Counter)
    two_tickers = []
    two_ticker_to_stock_id = {}
    
    for ticker in tickers_to_download:
        stock_id = ticker_to_stock_id[ticker]
        if ticker.endswith(".TW"):
            has_data = False
            if not df.empty:
                if len(tickers_to_download) == 1:
                    has_data = not df["Close"].dropna().empty
                else:
                    if ticker in df.columns.levels[0]:
                        has_data = not df[ticker]["Close"].dropna().empty
            
            if not has_data:
                two_ticker = f"{stock_id}.TWO"
                two_tickers.append(two_ticker)
                two_ticker_to_stock_id[two_ticker] = stock_id
                
    if two_tickers:
        print(f"Retrying {len(two_tickers)} OTC tickers with .TWO suffix...")
        try:
            df_two = yf.download(two_tickers, start=start_str, end=end_str, group_by="ticker")
            if not df_two.empty:
                if df.empty:
                    df = df_two
                else:
                    df = pd.concat([df, df_two], axis=1)
                for t_two, sid in two_ticker_to_stock_id.items():
                    stock_id_to_ticker[sid] = t_two
                    ticker_to_stock_id[t_two] = sid
        except Exception as e:
            print(f"yfinance download with .TWO suffix failed: {e}")

    # Map prices to stock data
    print("Mapping close prices to data structure...")
    for stock_id, sinfo in stock_data.items():
        ticker = stock_id_to_ticker.get(stock_id)
        history = sinfo["history"]
        for date in history.keys():
            if ticker:
                history[date]["close_price"] = get_close_price(df, ticker, date)

    # 4. Export as self-contained JS variable (avoids browser local CORS blocks)
    print("Writing JavaScript data file...")
    output_data = {
        "update_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "etf_info": etf_info,
        "stocks": stock_data
    }
    
    OUTPUT_JS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JS_PATH, "w", encoding="utf-8") as f:
        f.write("const ETF_DASHBOARD_DATA = ")
        json.dump(output_data, f, ensure_ascii=False, indent=2)
        f.write(";\n")
        
    print(f"Successfully generated dashboard data file at {OUTPUT_JS_PATH}")

if __name__ == "__main__":
    main()
