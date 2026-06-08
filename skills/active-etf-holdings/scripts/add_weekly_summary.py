import pandas as pd
import os
from datetime import datetime, timedelta

# Path to the Excel workbook containing daily ETF holdings sheets
EXCEL_PATH = r"i:/Mark/my-kb/wiki/金融投資/主動型ETF持股明細.xlsx"
OUTPUT_SHEET_NAME = "Weekly Summary"
DETAIL_SHEET_NAME = "Weekly Additions"

def parse_date(name: str):
    """Convert a sheet name like '20260522' to a datetime object.
    Returns None for non‑date sheets.
    """
    try:
        return datetime.strptime(name, "%Y%m%d")
    except Exception:
        return None

def load_date_sheets(path: str):
    """Load Excel file and return ExcelFile object and a mapping of date sheets.
    Handles the case where the file is corrupted or not a valid zip (e.g., after previous faulty edit).
    If the file cannot be read, a fresh workbook is created.
    """
    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        # If the file is not a valid Excel file, create a new one
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(path)
        xl = pd.ExcelFile(path)
    date_sheets = {}
    for name in xl.sheet_names:
        d = parse_date(name)
        if d:
            date_sheets[d] = name
    return xl, date_sheets

def get_weekly_comparison_dates(date_sheets: dict, end_date_str: str = None):
    """
    找出週比較的開始日期與結束日期。
    結束日期 (end_date)：
        若有指定 end_date_str，則為該日期。
        若無，則為 date_sheets 中的最新日期 (latest_date)。
    開始日期 (start_date)：
        結束日期所在週的週一前一週（即上週一至週五）的最大交易日。
        如果上週內無交易日，則 fallback 到小於結束日期所在週週一的最大交易日。
    """
    if not date_sheets:
        raise ValueError("No dated sheets found in the workbook.")
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y%m%d")
            if end_date not in date_sheets:
                # 尋找最接近且小於或等於指定日期的有資料日期
                valid_dates = [d for d in date_sheets if d <= end_date]
                if not valid_dates:
                    raise ValueError(f"No date sheets found on or before {end_date_str}")
                end_date = max(valid_dates)
        except Exception as e:
            raise ValueError(f"Invalid date format or date not found: {end_date_str}. Error: {e}")
    else:
        end_date = max(date_sheets.keys())
    
    # 結束日期所在週的週一
    monday = end_date - timedelta(days=end_date.weekday())
    
    # 上週的範圍 (Monday of last week to Friday of last week)
    last_week_start = monday - timedelta(days=7)
    last_week_end = monday - timedelta(days=3)  # Friday of last week is monday - 3 days
    
    # 找出上週範圍內的所有交易日
    last_week_dates = [d for d in date_sheets if last_week_start <= d <= last_week_end]
    
    if last_week_dates:
        start_date = max(last_week_dates)
    else:
        # Fallback: 小於結束週週一的最大日期
        older_dates = [d for d in date_sheets if d < monday]
        if older_dates:
            start_date = max(older_dates)
        else:
            # 如果連更早的都沒有，那就只能自己跟自己比了
            start_date = end_date
            
    return date_sheets[start_date], date_sheets[end_date]


def compute_weekly_changes(xl, start_name, end_name):
    """Compute total added and reduced 張數 for each ETF between two sheets."""
    df_start = xl.parse(start_name)
    df_end = xl.parse(end_name)
    # Standardize column names by stripping spaces
    df_start.columns = df_start.columns.str.strip()
    df_end.columns = df_end.columns.str.strip()
    # Ensure required columns exist
    for col in ['持有張數', 'ETF代號']:
        if col not in df_start.columns:
            raise KeyError(f"Column {col} not found in start sheet")
        if col not in df_end.columns:
            raise KeyError(f"Column {col} not found in end sheet")
    # Aggregate possible duplicate rows per ETF and holding code
    start_agg = df_start.groupby(['ETF代號', '持股代號'], as_index=False)['持有張數'].sum()
    end_agg = df_end.groupby(['ETF代號', '持股代號'], as_index=False)['持有張數'].sum()
    # Align indexes
    start = start_agg.set_index(['ETF代號', '持股代號'])['持有張數']
    end = end_agg.set_index(['ETF代號', '持股代號'])['持有張數']
    combined = pd.concat([start, end], axis=1, keys=['start', 'end']).fillna(0)
    combined['change'] = combined['end'] - combined['start']
    # Aggregate per ETF代號
    summary = combined.groupby(level=0)['change'].agg([
        lambda x: x[x > 0].sum(),   # total added
        lambda x: -x[x < 0].sum()   # total reduced (positive)
    ])
    summary.columns = ['總加碼張數', '總減碼張數']
    summary = summary.reset_index()
    # Add ETF名稱 if available
    if 'ETF名稱' in df_end.columns:
        etf_names = df_end.drop_duplicates('ETF代號')[['ETF代號', 'ETF名稱']].set_index('ETF代號')
        summary = summary.join(etf_names, on='ETF代號')
        summary = summary[['ETF代號', 'ETF名稱', '總加碼張數', '總減碼張數']]
    else:
        summary = summary[['ETF代號', '總加碼張數', '總減碼張數']]
    return summary

def compute_weekly_additions_detail(xl, start_name, end_name):
    """Return a DataFrame listing, for each ETF, the individual holdings that were added.
    Columns: ETF代號, ETF名稱 (if present), 持股代號, 持股名稱, 總加碼張數
    """
    df_start = xl.parse(start_name)
    df_end = xl.parse(end_name)
    df_start.columns = df_start.columns.str.strip()
    df_end.columns = df_end.columns.str.strip()
    # Ensure required columns exist
    for col in ['持有張數', 'ETF代號', '持股代號']:
        if col not in df_start.columns:
            raise KeyError(f"Column {col} not found in start sheet")
        if col not in df_end.columns:
            raise KeyError(f"Column {col} not found in end sheet")
    # Aggregate holdings for start and end sheets
    start_agg = df_start.groupby(['ETF代號', '持股代號'], as_index=False)['持有張數'].sum()
    end_agg = df_end.groupby(['ETF代號', '持股代號'], as_index=False)['持有張數'].sum()
    # Align on ETF代號 + 持股代號
    start = start_agg.set_index(['ETF代號', '持股代號'])['持有張數']
    end = end_agg.set_index(['ETF代號', '持股代號'])['持有張數']
    combined = pd.concat([start, end], axis=1, keys=['start', 'end']).fillna(0)
    combined['change'] = combined['end'] - combined['start']
    # Keep only positive additions
    added = combined[combined['change'] > 0].reset_index()
    if added.empty:
        # Return empty DataFrame with proper columns
        cols = ['ETF代號', 'ETF名稱', '持股代號', '持股名稱', '總加碼張數', '週原比例', '加碼後比例'] if 'ETF名稱' in df_end.columns else ['ETF代號', '持股代號', '持股名稱', '總加碼張數', '週原比例', '加碼後比例']
        return pd.DataFrame(columns=cols)
    # Map 持股名稱 from the latest sheet (end sheet)
    # Merge to get 持股名稱 without requiring unique index
    added = added.merge(df_end[['ETF代號', '持股代號', '持股名稱']].drop_duplicates(),
                        on=['ETF代號', '持股代號'], how='left')
    # Rename change column to 總加碼張數
    added = added.rename(columns={'change': '總加碼張數'})
    # --- Build 週原比例 and 加碼後比例 from 投資比例(%) ---
    ratio_col = '投資比例(%)'
    # Lookup from start sheet (週原比例)
    if ratio_col in df_start.columns:
        start_ratio = df_start[['ETF代號', '持股代號', ratio_col]].drop_duplicates()
        added = added.merge(start_ratio, on=['ETF代號', '持股代號'], how='left')
        added = added.rename(columns={ratio_col: '週原比例'})
    else:
        added['週原比例'] = None
    # Lookup from end sheet (加碼後比例)
    if ratio_col in df_end.columns:
        end_ratio = df_end[['ETF代號', '持股代號', ratio_col]].drop_duplicates()
        added = added.merge(end_ratio, on=['ETF代號', '持股代號'], how='left')
        added = added.rename(columns={ratio_col: '加碼後比例'})
    else:
        added['加碼後比例'] = None
    # 差異 = 加碼後比例 - 週原比例
    added['差異'] = pd.to_numeric(added['加碼後比例'], errors='coerce') - pd.to_numeric(added['週原比例'], errors='coerce')
    # Add ETF名稱 if present in end sheet
    if 'ETF名稱' in df_end.columns:
        etf_names = df_end[['ETF代號', 'ETF名稱']].drop_duplicates().set_index('ETF代號')
        added = added.join(etf_names, on='ETF代號')
        added = added[['ETF代號', 'ETF名稱', '持股代號', '持股名稱', '總加碼張數', '週原比例', '加碼後比例', '差異']]
    else:
        added = added[['ETF代號', '持股代號', '持股名稱', '總加碼張數', '週原比例', '加碼後比例', '差異']]
    return added

def compute_weekly_reductions_detail(xl, start_name, end_name):
    """Return a DataFrame listing, for each ETF, the individual holdings that were reduced.
    Columns: ETF代號, ETF名稱 (if present), 持股代號, 持股名稱, 總減碼張數
    """
    df_start = xl.parse(start_name)
    df_end = xl.parse(end_name)
    df_start.columns = df_start.columns.str.strip()
    df_end.columns = df_end.columns.str.strip()
    # Ensure required columns exist
    for col in ['持有張數', 'ETF代號', '持股代號']:
        if col not in df_start.columns:
            raise KeyError(f"Column {col} not found in start sheet")
        if col not in df_end.columns:
            raise KeyError(f"Column {col} not found in end sheet")
    # Aggregate holdings for start and end sheets
    start_agg = df_start.groupby(['ETF代號', '持股代號'], as_index=False)['持有張數'].sum()
    end_agg = df_end.groupby(['ETF代號', '持股代號'], as_index=False)['持有張數'].sum()
    start = start_agg.set_index(['ETF代號', '持股代號'])['持有張數']
    end = end_agg.set_index(['ETF代號', '持股代號'])['持有張數']
    combined = pd.concat([start, end], axis=1, keys=['start', 'end']).fillna(0)
    combined['change'] = combined['end'] - combined['start']
    # Keep only reductions (negative change)
    reduced = combined[combined['change'] < 0].reset_index()
    if reduced.empty:
        cols = ['ETF代號', 'ETF名稱', '持股代號', '持股名稱', '總減碼張數'] if 'ETF名稱' in df_end.columns else ['ETF代號', '持股代號', '持股名稱', '總減碼張數']
        return pd.DataFrame(columns=cols)
    # Map 持股名稱 from end sheet
    reduced = reduced.merge(df_end[['ETF代號', '持股代號', '持股名稱']].drop_duplicates(), on=['ETF代號', '持股代號'], how='left')
    # Fill missing 持股名稱 using start sheet as fallback (A+B approach)
    missing_mask = reduced['持股名稱'].isna()
    if missing_mask.any():
        # Create lookup from start sheet
        start_lookup = df_start[['ETF代號', '持股代號', '持股名稱']].drop_duplicates()
        reduced = reduced.merge(start_lookup, on=['ETF代號', '持股代號'], how='left', suffixes=('', '_start'))
        reduced.loc[missing_mask, '持股名稱'] = reduced.loc[missing_mask, '持股名稱_start']
        reduced = reduced.drop(columns=['持股名稱_start'])
    # Convert change to positive reduction amount
    reduced = reduced.rename(columns={'change': '總減碼張數'})
    reduced['總減碼張數'] = -reduced['總減碼張數']
    # --- Build 週原比例 and 減碼後比例 from 投資比例(%) ---
    ratio_col = '投資比例(%)'
    # Lookup from start sheet (週原比例)
    if ratio_col in df_start.columns:
        start_ratio = df_start[['ETF代號', '持股代號', ratio_col]].drop_duplicates()
        reduced = reduced.merge(start_ratio, on=['ETF代號', '持股代號'], how='left')
        reduced = reduced.rename(columns={ratio_col: '週原比例'})
    else:
        reduced['週原比例'] = None
    # Lookup from end sheet (減碼後比例)
    if ratio_col in df_end.columns:
        end_ratio = df_end[['ETF代號', '持股代號', ratio_col]].drop_duplicates()
        reduced = reduced.merge(end_ratio, on=['ETF代號', '持股代號'], how='left')
        reduced = reduced.rename(columns={ratio_col: '減碼後比例'})
        # 若全賣光，減碼後比例補 0
        reduced['減碼後比例'] = reduced['減碼後比例'].fillna(0)
    else:
        reduced['減碼後比例'] = None
    # 差異 = 減碼後比例 - 週原比例
    reduced['差異'] = pd.to_numeric(reduced['減碼後比例'], errors='coerce') - pd.to_numeric(reduced['週原比例'], errors='coerce')
    # Add ETF名稱 if present
    if 'ETF名稱' in df_end.columns:
        etf_names = df_end[['ETF代號', 'ETF名稱']].drop_duplicates().set_index('ETF代號')
        reduced = reduced.join(etf_names, on='ETF代號')
        reduced = reduced[['ETF代號', 'ETF名稱', '持股代號', '持股名稱', '總減碼張數', '週原比例', '減碼後比例', '差異']]
    else:
        reduced = reduced[['ETF代號', '持股代號', '持股名稱', '總減碼張數', '週原比例', '減碼後比例', '差異']]
    return reduced

def move_sheet_to_first(path: str, sheet_name: str):
    """Reorder the workbook so that *sheet_name* becomes the first sheet.
    Uses openpyxl which is compatible with the 'openpyxl' engine used by
    pandas.ExcelWriter.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[sheet_name]
    # openpyxl's move_sheet moves relative to current position; calculate offset
    current_index = wb.sheetnames.index(sheet_name)
    if current_index != 0:
        wb.move_sheet(ws, offset=-current_index)
    wb.save(path)

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Update weekly additions and reductions summary in Excel.")
    parser.add_argument("--date", type=str, help="Specify the end date in YYYYMMDD format for comparison.")
    parser.add_argument("--force", action="store_true", help="Force update even if today is not a weekend day.")
    args = parser.parse_args()
    
    # 判斷是否應該執行
    # 預設僅在週五、六、日 (weekday 4, 5, 6) 執行
    today_weekday = datetime.now().weekday()
    is_weekend = today_weekday in [4, 5, 6]
    
    if not is_weekend and not args.force and not args.date:
        print(f"Today is weekday {today_weekday + 1} (not Friday, Saturday, or Sunday).")
        print("Weekly summary update is skipped to avoid overwriting previous weekend data.")
        print("To force update, use --force or specify a date with --date YYYYMMDD.")
        return

    xl, date_sheets = load_date_sheets(EXCEL_PATH)
    try:
        start_name, end_name = get_weekly_comparison_dates(date_sheets, args.date)
    except Exception as e:
        print(f"Error determining comparison dates: {e}")
        return

    if start_name == end_name:
        print(f"Warning: Start date and End date are the same ({end_name}). No changes will be detected.")
        
    print(f"Comparing end date '{end_name}' with start date '{start_name}' (last week's end).")
    
    additions_df = compute_weekly_additions_detail(xl, start_name, end_name)
    reductions_df = compute_weekly_reductions_detail(xl, start_name, end_name)


    # Write Weekly Additions and Weekly Reductions sheets
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        additions_df.to_excel(writer, sheet_name=DETAIL_SHEET_NAME, index=False)
        reductions_df.to_excel(writer, sheet_name='Weekly Reductions', index=False)
        
        # Apply formatting: freeze panes and autofit column width
        for sheet_name in [DETAIL_SHEET_NAME, 'Weekly Reductions']:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                column_letter = column_cells[0].column_letter
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 80)
    # Apply red font to holdings that are fully sold out (final holding = 0)
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    # Reload Excel file to safely read end sheet after writing
    xl_new = pd.ExcelFile(EXCEL_PATH)
    wb = load_workbook(EXCEL_PATH)
    # Load start and end sheets
    df_start = xl_new.parse(start_name)
    df_start.columns = df_start.columns.str.strip()
    df_end = xl_new.parse(end_name)
    df_end.columns = df_end.columns.str.strip()
    # Build a lookup dictionary for end holdings (ETF代號, 持股代號) -> 持有張數
    end_lookup = {
        (str(row['ETF代號']).strip(), str(row['持股代號']).strip()): row['持有張數']
        for _, row in df_end.iterrows()
    }
    # Determine zero or missing holdings
    zero_holdings = set()
    for _, row in df_start.iterrows():
        key = (str(row['ETF代號']).strip(), str(row['持股代號']).strip())
        qty = end_lookup.get(key, None)
        if qty is None or qty == 0:
            zero_holdings.add(key)
    ws_red = wb['Weekly Reductions']
    # Identify columns indices for matching
    header = [cell.value for cell in next(ws_red.iter_rows(min_row=1, max_row=1))]
    etf_col = header.index('ETF代號') + 1
    code_col = header.index('持股代號') + 1
    for row in ws_red.iter_rows(min_row=2):
        etf = row[etf_col - 1].value
        code = row[code_col - 1].value
        # Convert to string for reliable matching
        etf_str = str(etf).strip() if etf is not None else None
        code_str = str(code).strip() if code is not None else None
        if (etf_str, code_str) in zero_holdings:
            for cell in row:
                cell.font = Font(color="FF0000")
    # Apply blue font to Weekly Additions if "週原比例" is 0 or missing
    if DETAIL_SHEET_NAME in wb.sheetnames:
        ws_add = wb[DETAIL_SHEET_NAME]
        header_add = [cell.value for cell in next(ws_add.iter_rows(min_row=1, max_row=1))]
        if '週原比例' in header_add:
            orig_ratio_col = header_add.index('週原比例') + 1
            for row in ws_add.iter_rows(min_row=2):
                orig_ratio = row[orig_ratio_col - 1].value
                if orig_ratio is None or str(orig_ratio).strip() == '' or orig_ratio == 0:
                    for cell in row:
                        cell.font = Font(color="0000FF")
    # Remove Weekly Summary sheet if it exists
    if OUTPUT_SHEET_NAME in wb.sheetnames:
        ws = wb[OUTPUT_SHEET_NAME]
        wb.remove(ws)
    wb.save(EXCEL_PATH)
    # Ensure sheet order: Weekly Additions first, Weekly Reductions second
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    # Move Weekly Additions to first (already done by move_sheet_to_first, but ensure again)
    if DETAIL_SHEET_NAME in wb.sheetnames:
        ws_add = wb[DETAIL_SHEET_NAME]
        cur_idx = wb.sheetnames.index(DETAIL_SHEET_NAME)
        if cur_idx != 0:
            wb.move_sheet(ws_add, offset=-cur_idx)
    # Move Weekly Reductions to second position
    red_sheet = 'Weekly Reductions'
    if red_sheet in wb.sheetnames:
        ws_red = wb[red_sheet]
        cur_idx = wb.sheetnames.index(red_sheet)
        # Desired position is 1 (second tab)
        desired_idx = 1
        offset = desired_idx - cur_idx
        if offset != 0:
            wb.move_sheet(ws_red, offset=offset)
    wb.save(EXCEL_PATH)
    wb.close()

if __name__ == "__main__":
    main()
