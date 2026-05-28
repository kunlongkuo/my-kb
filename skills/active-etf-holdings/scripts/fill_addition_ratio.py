#!/usr/bin/env python
"""Fill addition ratio values in 加碼張數 sheet of the ETF holdings Excel.

The existing sheet may contain formulas for 加碼張數 and 加碼比例.
This script replaces those formulas with static numeric values:
    加碼張數 = 加碼後張數 - 原張數
    加碼比例 = IF(原張數=0, 0, (加碼後張數-原張數)/原張數)
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

def main() -> None:
    excel_path = Path(r"i:\\Mark\\my-kb\\wiki\\金融投資\\主動型ETF持股明細.xlsx")
    if not excel_path.is_file():
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    wb = load_workbook(excel_path)
    sheet_name = "加碼張數"
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found in workbook.")
        sys.exit(1)

    ws = wb[sheet_name]
    # Assume first row is header
    header = [cell.value for cell in ws[1]]
    try:
        idx_original = header.index("原張數") + 1  # 1‑based column index
        idx_new = header.index("加碼後張數") + 1
    except ValueError:
        print("Required columns 原張數 or 加碼後張數 not found in header.")
        sys.exit(1)

    # Target columns for static values (originally formula columns)
    idx_diff = idx_new + 1   # 加碼張數 column
    idx_pct = idx_new + 2    # 加碼比例 column

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        original = row[idx_original - 1].value
        new = row[idx_new - 1].value
        # Normalise None / empty to 0 for arithmetic
        original_val = float(original) if original not in (None, "") else 0.0
        new_val = float(new) if new not in (None, "") else 0.0
        if original_val == 0:
            diff = 0.0
            pct = 0.0
        else:
            diff = new_val - original_val
            pct = diff / original_val
        ws.cell(row=row[0].row, column=idx_diff, value=diff)
        ws.cell(row=row[0].row, column=idx_pct, value=pct)
        # Optional: set number formats for readability
        ws.cell(row=row[0].row, column=idx_diff).number_format = '#,##0.00'
        ws.cell(row=row[0].row, column=idx_pct).number_format = '0.00%'

    wb.save(excel_path)
    print(f"Values filled in sheet '{sheet_name}' of {excel_path}")

if __name__ == "__main__":
    main()
