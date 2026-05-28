#!/usr/bin/env python
"""Inspect the '加碼張數' sheet in the ETF holdings Excel file.
Print column headers and first few data rows (values only).
"""

from pathlib import Path
import openpyxl

def main():
    excel_path = Path(r"i:/Mark/my-kb/wiki/金融投資/主動型ETF持股明細.xlsx")
    if not excel_path.is_file():
        print(f"File not found: {excel_path}")
        return
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = "加碼張數"
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found.")
        return
    ws = wb[sheet_name]
    # Header
    header = [cell.value for cell in ws[1]]
    print("Header:", header)
    # Print first 5 data rows
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        print(f"Row {i}:", row)

if __name__ == "__main__":
    main()
