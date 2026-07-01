#!/usr/bin/env python
"""
draw_holdings_charts.py

讀取「每日個股合計」工作表，找出最新一日的加碼/減碼 Top 10 個股，
利用 matplotlib 繪製精美的深色系（Dark Mode）橫向長條圖，
保存至 wiki/金融投資/images/ 目錄，並在 主動型ETF持股變動.md 中動態嵌入圖表。
"""

from __future__ import annotations

import os
import re
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# 註冊中文字型（防止 Windows 系統下中文顯示亂碼）
# ─────────────────────────────────────────────────────────────────────────────
font_path = "C:\\Windows\\Fonts\\msjh.ttc"
if not os.path.exists(font_path):
    font_path = "C:\\Windows\\Fonts\\msjh.ttf"

if os.path.exists(font_path):
    try:
        fe = fm.FontEntry(fname=font_path, name="Microsoft JhengHei")
        fm.fontManager.ttflist.insert(0, fe)
        plt.rcParams['font.family'] = ['Microsoft JhengHei']
        print("已成功註冊並設定微軟正黑體。")
    except Exception as e:
        print(f"註冊微軟正黑體失敗: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 路徑定義
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
XLSX_PATH = BASE_DIR / "wiki" / "金融投資" / "主動型ETF持股明細.xlsx"
CHANGES_MD_PATH = BASE_DIR / "wiki" / "金融投資" / "主動型ETF持股變動.md"
IMAGES_DIR = BASE_DIR / "wiki" / "金融投資" / "images"

# ─────────────────────────────────────────────────────────────────────────────
# 讀取 Excel 數據
# ─────────────────────────────────────────────────────────────────────────────
def load_daily_totals() -> list[dict]:
    if not XLSX_PATH.exists():
        print(f"找不到 Excel 檔案: {XLSX_PATH}")
        return []
    
    print(f"讀取 Excel: {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if "每日個股合計" not in wb.sheetnames:
        print("Excel 中找不到 '每日個股合計' 分頁。")
        return []
    
    ws = wb["每日個股合計"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    
    headers = [str(h or "").strip() for h in rows[0]]
    
    # 取得必要欄位的索引
    try:
        idx_date = headers.index("日期")
        idx_code = headers.index("持股代號")
        idx_name = headers.index("持股名稱")
        idx_delta = headers.index("張數增減")
        idx_etf_count = headers.index("涵蓋ETF數")
        idx_total_lots = headers.index("當日總張數")
    except ValueError as e:
        print(f"工作表欄位不完整: {e}")
        return []
        
    data = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        date_val = str(r[idx_date] or "").strip()
        code_val = str(r[idx_code] or "").strip()
        name_val = str(r[idx_name] or "").strip()
        
        delta_val = r[idx_delta]
        etf_count_val = r[idx_etf_count]
        total_lots_val = r[idx_total_lots]
        
        # 轉換數值
        try:
            delta = float(delta_val) if delta_val is not None else 0.0
        except (ValueError, TypeError):
            delta = 0.0
            
        try:
            etf_count = int(etf_count_val) if etf_count_val is not None else 0
        except (ValueError, TypeError):
            etf_count = 0
            
        try:
            total_lots = float(total_lots_val) if total_lots_val is not None else 0.0
        except (ValueError, TypeError):
            total_lots = 0.0
            
        data.append({
            "date": date_val,
            "code": code_val,
            "name": name_val,
            "delta": delta,
            "etf_count": etf_count,
            "total_lots": total_lots
        })
    return data

# ─────────────────────────────────────────────────────────────────────────────
# 繪製圖表
# ─────────────────────────────────────────────────────────────────────────────
def draw_chart(
    items: list[dict], 
    is_addition: bool, 
    date_str: str, 
    output_filename: str,
    fixed_filename: str
):
    # 排序與準備數據
    if is_addition:
        title = f"今日主動型 ETF 加碼排行 Top 10 ({date_str})"
        # 降序排序，取前 10
        sorted_items = sorted([x for x in items if x["delta"] > 0], key=lambda x: x["delta"], reverse=True)[:10]
        color = "#10B981"  # Emerald 500
        x_label = "加碼張數"
    else:
        title = f"今日主動型 ETF 減碼排行 Top 10 ({date_str})"
        # 升序排序（負得最多），取前 10，並將 delta 轉為正數以利視覺長條延伸
        sorted_items = sorted([x for x in items if x["delta"] < 0], key=lambda x: x["delta"])[:10]
        for x in sorted_items:
            x["display_delta"] = abs(x["delta"])
        color = "#EF4444"  # Red 500
        x_label = "減碼張數"
        
    if not sorted_items:
        print(f"無相關資料，跳過繪製 {title}")
        return False
        
    # 反轉順序以利在橫條圖中由上往下（第一名在最上面）
    sorted_items = sorted_items[::-1]
    
    # 準備繪圖數據
    labels = []
    values = []
    for x in sorted_items:
        label = f"{x['code']} {x['name']}" if x['code'] else x['name']
        labels.append(label)
        values.append(x["display_delta"] if not is_addition else x["delta"])
        
    # 設定 Dark Mode 樣式
    plt.style.use('dark_background')
    fig, ax = plt.subplots(figsize=(6.5, 4.5), facecolor="#111827")
    ax.set_facecolor("#111827")
    
    # 畫橫向長條圖
    bars = ax.barh(labels, values, color=color, edgecolor="none", height=0.6, alpha=0.9)
    
    # 美化坐標軸與網格
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#4B5563')
    ax.spines['left'].set_color('#4B5563')
    
    ax.tick_params(colors='#F3F4F6', labelsize=10)
    ax.xaxis.grid(True, linestyle='--', alpha=0.2, color='#9CA3AF')
    ax.set_axisbelow(True)
    
    # 設定支援負號
    plt.rcParams['axes.unicode_minus'] = False
    
    ax.set_title(title, fontsize=12, fontweight='bold', color='#F3F4F6', pad=15)
    ax.set_xlabel(x_label, fontsize=10, color='#9CA3AF', labelpad=8)
    
    # 於長條右側標註數值
    max_val = max(values) if values else 1
    for bar in bars:
        width = bar.get_width()
        ax.text(
            width + (max_val * 0.015), 
            bar.get_y() + bar.get_height()/2, 
            f"{width:,.1f} 張", 
            va='center', 
            ha='left', 
            fontsize=9.5, 
            color='#F3F4F6', 
            fontweight='bold'
        )
        
    # 調整間距，避免文字被裁切
    plt.tight_layout()
    
    # 確保資料夾存在
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    
    # 儲存檔案
    p_dated = IMAGES_DIR / output_filename
    p_fixed = IMAGES_DIR / fixed_filename
    
    plt.savefig(p_dated, dpi=150, facecolor="#111827", edgecolor="none")
    plt.savefig(p_fixed, dpi=150, facecolor="#111827", edgecolor="none")
    plt.close()
    
    print(f"成功儲存圖表: {p_dated} & {p_fixed}")
    return True

# ─────────────────────────────────────────────────────────────────────────────
# 更新 Markdown 文件
# ─────────────────────────────────────────────────────────────────────────────
def update_markdown_embed(date_yyyymmdd: str):
    if not CHANGES_MD_PATH.exists():
        print(f"找不到變動分析 MD 檔案: {CHANGES_MD_PATH}")
        return
        
    print(f"更新 Markdown 嵌入: {CHANGES_MD_PATH}")
    text = CHANGES_MD_PATH.read_text(encoding="utf-8")
    
    # 準備插入的圖片並排 Markdown 語法
    embed_content = (
        f"\n## 今日加減碼視覺化排行\n\n"
        f"| 今日加碼 Top 10 | 今日減碼 Top 10 |\n"
        f"| :---: | :---: |\n"
        f"| ![今日加碼 Top 10](images/{date_yyyymmdd}_additions.png) | ![今日減碼 Top 10](images/{date_yyyymmdd}_reductions.png) |\n"
    )
    
    # 檢查是否已存在該區段，若存在則替換，否則插入到比較區間下方
    section_pattern = r"\n## 今日加減碼視覺化排行\n.*?(\n## [0-9]{5}[A-Z]|\Z)"
    if re.search(r"## 今日加減碼視覺化排行", text):
        # 替換舊有的區段
        text = re.sub(
            r"\n## 今日加減碼視覺化排行\n.*?(?=\n## [0-9]{5}[A-Z]|\Z)", 
            embed_content, 
            text, 
            flags=re.S
        )
        print("已更新既有的視覺化排行區段。")
    else:
        # 插入在比較區間：`YYYYMMDD` → `YYYYMMDD` 下方
        compare_match = re.search(r"比較區間：`[0-9]+` → `[0-9]+`.*?\n", text)
        if compare_match:
            insert_pos = compare_match.end()
            text = text[:insert_pos] + embed_content + text[insert_pos:]
            print("已插入新的視覺化排行區段。")
        else:
            # 若無比較區間標記，則直接插在第 4 行後面
            lines = text.splitlines()
            if len(lines) >= 4:
                lines.insert(4, embed_content)
                text = "\n".join(lines)
                print("插在文件開頭部分。")
                
    CHANGES_MD_PATH.write_text(text, encoding="utf-8")
    print("Markdown 檔案寫入成功。")

# ─────────────────────────────────────────────────────────────────────────────
# 主程式
# ─────────────────────────────────────────────────────────────────────────────
def main():
    data = load_daily_totals()
    if not data:
        print("無可用數據。")
        return
        
    # 找出最新日期
    dates = sorted(list(set(x["date"] for x in data)))
    if not dates:
        print("數據中找不到任何日期。")
        return
        
    latest_date_slash = dates[-1]  # YYYY/MM/DD 格式
    date_yyyymmdd = latest_date_slash.replace("/", "")  # YYYYMMDD 格式
    
    print(f"最新資料日期: {latest_date_slash} ({date_yyyymmdd})")
    
    # 篩選出最新日期的列
    latest_data = [x for x in data if x["date"] == latest_date_slash]
    
    # 畫加碼排行
    add_ok = draw_chart(
        latest_data, 
        is_addition=True, 
        date_str=latest_date_slash,
        output_filename=f"{date_yyyymmdd}_additions.png",
        fixed_filename="active_etf_top_additions.png"
    )
    
    # 畫減碼排行
    dec_ok = draw_chart(
        latest_data, 
        is_addition=False, 
        date_str=latest_date_slash,
        output_filename=f"{date_yyyymmdd}_reductions.png",
        fixed_filename="active_etf_top_reductions.png"
    )
    
    if add_ok or dec_ok:
        update_markdown_embed(date_yyyymmdd)

if __name__ == "__main__":
    main()
